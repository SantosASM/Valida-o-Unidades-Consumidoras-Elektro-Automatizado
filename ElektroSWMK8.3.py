
# -*- coding: utf-8 -*-
"""
ElektroSWMKEight.py
Automação Elektro (Neoenergia) — Segunda Via via WhatsApp Web

MELHORIAS EM RELAÇÃO AO MK7:
  Bug 1 → OFFSET de linha: wb_row era calculado de df_proc.reset_index (sempre 0..N)
           mesmo quando LINHA_INICIAL > 1, gravando resultados nas linhas erradas.
           Corrigido: wb_row = ini + df_idx + 2 (referenciado ao df inteiro).

  Bug 2 → Arquivo TMP aberto no Excel bloqueava a escrita (PermissionError silenciado).
           O script salvava em arquivo com timestamp sem avisar claramente.
           Corrigido: aviso explícito + estratégia robusta de nomes alternativos.

  Bug 3 → Arquivo de saída gerado do zero, sem depender de arquivo de entrada.
           O workbook agora é criado internamente; não precisa existir previamente.

  Bug 4 → df_proc era criado ANTES de criar_workbook_resultado adicionar as colunas
           obs/validacao ao df. Agora as colunas são garantidas ANTES de qualquer split.

  Bug 5 → ws.cell() não tinha try/except; erros silenciosos causavam linhas em branco.
           Corrigido: bloco try/except com log de erro por linha.

MK8 — Novidades (mantidas nesta revisão):
  · Leitura de mensagens via JS (E1–E6) com detecção de posição para enviadas/recebidas.
  · processar_uc_unica com texto_uc= para evitar stale element do DOM.
  · Detecção de "sem fatura" em camadas extras antes/durante P5→P6/P7.
  · Colunas extras na planilha: Ativas, Inativas, Fatura baixada.
  · tempo de execução adicionado ao resumo enviado ao grupo GOC.
  · Leitura de CSV tentando múltiplos separadores e codificações.

SINCRONIZAÇÃO COM MK7.2 (esta revisão):
  A lógica de CLASSIFICAÇÃO DE ERROS foi revertida para ser idêntica à do MK7.2,
  pois o MK8 estava classificando situações que não são erro como erro/retry:
    · ERROS_PARA_RETRY / eh_erro_de_retry() substituem FALHAS_DEFINITIVAS /
      eh_falha_definitiva() — o padrão volta a ser "falha definitiva" por default,
      com retry apenas para a lista explícita de erros transitórios.
    · detectar_erro_bot() voltou a incluir o ramo "melhorar nosso serviço".
    · _eh_modo_avaliacao() e seus pontos de chamada em processar_fluxo (P2/P3/P4
      e fim de P6/P7) foram removidos — esse cenário volta a ser tratado só por
      detectar_erro_bot(), como no MK7.2.
    · processar_uc_unica / processar_lista_uc voltam a retornar "UC não confere..."
      e "mais de uma UC..." (em vez de "Inativa...") quando a UC não bate com a
      apresentada pelo bot.
    · processar_linha() usa o mesmo encadeamento elif do MK7.2 (Erro CPF / mais de
      uma UC / eh_erro_de_retry / else falha).
  Os mecanismos de LEITURA do DOM (seletores, _ler_msgs_js, aguardar_bot com
  checagem multi-mensagem, FRASES_ENCERRAMENTO expandida, validar_saudacao_neoenergia
  com espera extra de 5s, leitura de CSV multi-encoding) permanecem os do MK8 —
  eles não fazem parte da lógica de classificação de erro e continuam a melhorar
  a robustez de detecção independentemente da classificação final.
"""

import re
import sys
import time
import unicodedata
from datetime import datetime, date as _date, time as _time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
from playwright.sync_api import sync_playwright

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────────────────────────
PASTA_PROJETO = Path(r"C:\Users\Santos\Desktop\Elektro")
ARQUIVO_ENTRADA = PASTA_PROJETO / "Elektro_1.csv"
PASTA_RESULTADO = PASTA_PROJETO / "Elektro_Resultado"
ARQUIVO_RESULTADO_TMP = PASTA_RESULTADO / "Elektro_resultado_em_andamento.xlsx"
PASTA_PERFIL = PASTA_PROJETO / ".perfil_whatsapp"
PASTA_FATURAS = PASTA_PROJETO / "Faturas"

CONTATO_BOT = "Elektro whats validação"
GRUPO_NOTIFICACAO = "GOC"

TIMEOUT_BOT = 120
TIMEOUT_REENVIO = 30
TIMEOUT_POS_CLIQUE_PAINEL = 25
PAUSA_ENTRE_ITENS = 5

LINHA_INICIAL = 1
LINHA_FINAL = None

ERROS_PARA_RETRY = {
    "ERRO: bot não respondeu",
    "ERRO: bot não pediu CPF",
    "ERRO: bot não confirmou UC",
    "ERRO: bot perguntou 'posso te ajudar' após confirmar UC",
    "Erro: fluxo interrompido",
    "Erro: saudação Neoenergia não recebida",
    "ERRO: painel de UC não respondeu",
    "Erro ao baixar fatura",
    # Avaliação detectada ANTES de confirmar a UC (em P2/P3/P4): o fluxo quebrou
    # antes de validar. Mandar 'Sair' reinicia o bot; tentamos a UC de novo.
    # (Quando a avaliação ocorre DEPOIS de confirmar a UC, o fluxo já retorna
    #  "Ativa - sem segunda via para emissão" e nem chega aqui.)
    "bot pediu avaliação após encerramento",
}

SAUDACAO_NEOENERGIA = "assistente virtual da neoenergia elektro"
FRASE_MENU_SERVICOS = "escolha o serviço para o qual"
FRASE_PIX = "para pagar com pix"
FRASE_DIVIDAS = "consultar se existem dívidas"
FRASE_DIVIDAS_ALT = "consultar se existem dividas"
FRASE_INFO_PAGAMENTO = "informações para pagamento da fatura"
FRASE_INFO_PAGAMENTO_ALT = "informacoes para pagamento da fatura"
FRASE_ESCOLHA_FATURA = "escolha a fatura que gostaria de emitir"
FRASE_ESCOLHA_FATURA_ALT = "escolha a fatura que gostaria"
FRASE_FALTOU_NUMERO = "faltou algum número"
FRASE_FALTOU_NUMERO_ALT = "faltou algum numero"
FRASE_FALTOU_NUMERO_POXA = "poxa..."
TIMEOUT_ESPERA_PDF_POS_INFO = 20

FRASES_ENCERRAMENTO = [
    # As DUAS mensagens finais oficiais do bot. O emoji ⚡ pode vir como caractere
    # OU ficar oculto quando renderizado como <img alt="">, por isso há fragmentos
    # de texto que cobrem ambos os casos.
    #
    # Final 1: "Tudo bem! Se precisar de mais alguma coisa, estou por aqui.
    #           Conte sempre com a gente! 😉👋"
    "tudo bem! se precisar", "estou por aqui", "conte sempre com a gente",
    #
    # Final 2: "⚡ # Dica de Segurança: Fique longe de fios e postes.
    #           Distância mínima de segurança: 2,5m do fio do poste."
    "dica de segurança", "dica de seguranca",
    "# dica de seguran",   # começa com # porque o emoji ⚡ vira <img>
    "fique longe de fios", "fique longe de fios e postes",
    "distância mínima", "distancia minima", "distancia minima de seguranca",
    "fio do poste", "2,5m do fio",
]

SEL_MSG_RECEBIDA = "div.message-in"  # fallback legado; leitura real feita via JS

# JavaScript que lê mensagens RECEBIDAS direto do DOM do WhatsApp Web.
# 6 estratégias em cascata para acomodar qualquer versão do WhatsApp Web.
# Estratégia 6 (última) trabalha diretamente com span.copyable-text,
# independente do container pai — mais robusta quando classes são ofuscadas.
_JS_LER_MSGS = """
() => {
    const main = document.querySelector('#main');
    if (!main) return {count: 0, ultimo: '', todos: [], debug: 'sem #main'};

    // Seletores de ícones de entrega presentes apenas em mensagens ENVIADAS
    const SEL_SAIDA = '[data-icon="msg-check"],[data-icon="msg-dblcheck"],' +
                      '[data-icon="msg-dblcheck-ack"],[data-icon="msg-time"]';

    // Cache do rect do painel principal (chamado uma vez por execução do JS)
    const mainRect = main.getBoundingClientRect();

    // Retorna true se a mensagem é ENVIADA pelo usuário.
    // Método 1: delivery icon dentro do elemento (mais rápido).
    // Método 2: posição do bubble — mensagens enviadas ficam no lado DIREITO do painel.
    //   O centro do span de texto > 50% da largura do painel indica lado direito = enviado.
    //   Isso é robusto mesmo quando o data-icon está fora do [role="row"].
    const ehSaida = (el) => {
        if (el.querySelector(SEL_SAIDA)) return true;
        // Fallback por posição: calcula centro do bubble relativo ao painel
        const bubble = el.querySelector('span.copyable-text, span.selectable-text');
        if (bubble && mainRect.width > 0) {
            try {
                const bRect = bubble.getBoundingClientRect();
                const relCenter = ((bRect.left + bRect.right) / 2 - mainRect.left) / mainRect.width;
                if (relCenter > 0.5) return true; // centro no lado direito = enviado
            } catch(e) {}
        }
        return false;
    };

    // Extrai texto mais externo da mensagem: procura span.copyable-text / span.selectable-text
    // não aninhado em outro, ou cai no innerText completo do elemento.
    const getTexto = (el) => {
        const sel = 'span.copyable-text, span.selectable-text, span[class*="copyable"]';
        const spans = Array.from(el.querySelectorAll(sel));
        for (const s of spans) {
            const pai = s.parentElement;
            if (!pai || !pai.closest(sel)) {
                const txt = (s.innerText || '').trim();
                if (txt.length > 0) return txt;
            }
        }
        return (el.innerText || '').trim();
    };

    const temTexto = (el) => !!el.querySelector(
        'span.copyable-text, span.selectable-text, span[class*="copyable"]'
    );

    let msgs = [], estrategia = 0;

    // E1: classe clássica
    if (!msgs.length) {
        msgs = Array.from(main.querySelectorAll('div.message-in'));
        if (msgs.length) estrategia = 1;
    }
    // E2: classe parcial
    if (!msgs.length) {
        msgs = Array.from(main.querySelectorAll('div[class*="message-in"]'));
        if (msgs.length) estrategia = 2;
    }
    // E3: role=row sem ícone de entrega
    if (!msgs.length) {
        msgs = Array.from(main.querySelectorAll('[role="row"]'))
               .filter(r => !ehSaida(r) && temTexto(r));
        if (msgs.length) estrategia = 3;
    }
    // E4: role=listitem sem ícone de entrega
    if (!msgs.length) {
        msgs = Array.from(main.querySelectorAll('[role="listitem"]'))
               .filter(r => !ehSaida(r) && temTexto(r));
        if (msgs.length) estrategia = 4;
    }
    // E5: sobe a partir de span.copyable-text até achar container com role ou li
    if (!msgs.length) {
        const visto = new Set();
        const sel = 'span.copyable-text, span.selectable-text';
        Array.from(main.querySelectorAll(sel)).forEach(s => {
            if (s.parentElement && s.parentElement.closest(sel)) return; // aninhado
            let node = s.parentElement, achou = false;
            for (let i = 0; i < 12 && node && node !== main; i++) {
                const r = node.getAttribute && node.getAttribute('role');
                if (r === 'row' || r === 'listitem' || node.tagName === 'LI') { achou = true; break; }
                node = node.parentElement;
            }
            // Só usa o container se encontrou um com role explícito
            if (!achou || !node || node === main || visto.has(node)) return;
            if (!ehSaida(node)) { visto.add(node); msgs.push(node); }
        });
        if (msgs.length) estrategia = 5;
    }
    // E6 (último recurso): extrai textos diretamente dos spans, sem container
    // Útil quando a estrutura DOM é completamente diferente do esperado.
    if (!msgs.length) {
        const textos = [];
        const sel = 'span.copyable-text, span.selectable-text, span[class*="copyable"]';
        Array.from(main.querySelectorAll(sel)).forEach(s => {
            if (s.parentElement && s.parentElement.closest(sel)) return; // aninhado
            // Método 1: ícone de entrega em algum ancestral
            let isOut = false, node = s.parentElement;
            for (let i = 0; i < 14 && node && node !== main; i++) {
                if (node.querySelector && node.querySelector(SEL_SAIDA)) { isOut = true; break; }
                node = node.parentElement;
            }
            // Método 2: posição do span (centro > 50% do painel = lado direito = enviado)
            if (!isOut && mainRect.width > 0) {
                try {
                    const sRect = s.getBoundingClientRect();
                    const relCenter = ((sRect.left + sRect.right) / 2 - mainRect.left) / mainRect.width;
                    if (relCenter > 0.5) isOut = true;
                } catch(e) {}
            }
            if (!isOut) {
                const txt = (s.innerText || '').trim();
                if (txt.length > 0) textos.push(txt);
            }
        });
        if (textos.length) {
            return {count: textos.length, ultimo: textos[textos.length - 1],
                    todos: textos.slice(-10), debug: 'E6-spans'};
        }
    }

    if (!msgs.length) return {count: 0, ultimo: '', todos: [], debug: 'E0-falhou'};

    const ultimo = getTexto(msgs[msgs.length - 1]);
    const todos = msgs.slice(-10).map(getTexto).filter(t => t.length > 0);
    return {count: msgs.length, ultimo: ultimo, todos: todos, debug: 'E' + estrategia};
}
"""

SELETORES_CAIXA_BUSCA = [
    'input[role="textbox"][data-tab="3"]',
    'input[aria-label="Pesquisar ou começar uma nova conversa"]',
    'input[aria-label="Search or start new chat"]',
    'div[contenteditable="true"][data-tab="3"]',
    'div[contenteditable="true"][role="textbox"]',
    'div[aria-label="Caixa de texto de pesquisa"]',
    'div[aria-label="Search input textbox"]',
    'div[title="Caixa de texto de pesquisa"]',
]

SELETORES_CAIXA_MSG = [
    'footer div[contenteditable="true"]',
    'div[contenteditable="true"][data-tab="10"]',
    'div[aria-label="Digite uma mensagem"]',
    'div[aria-label="Type a message"]',
]

# Nomes canônicos das colunas de resultado (sem acento, minúsculas)
# A planilha de saída agora tem apenas duas colunas de resultado:
#   · "Resultado" — registro final único por UC (última tentativa quando há 2)
#   · "Busca"     — preenchida APENAS quando um localizador não foi encontrado
#                   (ex.: "Não encontrei a caixa de mensagem"), com o que foi buscado
COL_RESULTADO = "resultado"
COL_BUSCA     = "busca"

# Valores canônicos da coluna "Resultado".
RES_ATIVA          = "Ativa"
RES_ATIVA_FATURA   = "Ativa e Fatura baixada"
RES_INATIVA        = "Inativa"
# Para linhas puladas, o próprio motivo do pulo vai na coluna Resultado.


# ──────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS
# ──────────────────────────────────────────────────────────────────────────────

def _remover_acentos(texto: str) -> str:
    """Remove acentos e normaliza para ASCII lowercase."""
    return unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("ascii").lower()


def digitos(valor):
    return re.sub(r"\D", "", str(valor if valor is not None else ""))


def parse_data(valor):
    if valor is None:
        return None
    if isinstance(valor, _time):
        return None
    if isinstance(valor, (datetime, _date)):
        return valor.strftime("%d/%m/%Y")
    s = str(valor).strip()
    if s.lower() in ("", "0", "undefined", "none", "nan", "nat"):
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", s.strip())
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    return None


def eh_erro_de_retry(obs):
    obs_s = str(obs or "").strip()
    return any(obs_s.startswith(e) for e in ERROS_PARA_RETRY)


def formatar_duracao(segundos: float) -> str:
    """Formata segundos em string legível: 1h 23min 45s."""
    segundos = int(segundos)
    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    segs = segundos % 60
    partes = []
    if horas:
        partes.append(f"{horas}h")
    if minutos:
        partes.append(f"{minutos}min")
    partes.append(f"{segs}s")
    return " ".join(partes)


# ──────────────────────────────────────────────────────────────────────────────
# LÓGICA DE DETECÇÃO / VERIFICAÇÃO DO BOT
# ──────────────────────────────────────────────────────────────────────────────

def uc_bate(uc_arquivo, texto_item):
    uc_arq = digitos(uc_arquivo)
    if len(uc_arq) < 4:
        return False
    primeiros_arq = uc_arq[:2]
    ultimos_arq = uc_arq[-2:]
    # UC curta (ex.: '6467'): o bot mascara mostrando outro sufixo (ex.: '64** ... 224'),
    # então os 2 últimos do arquivo não correspondem ao sufixo do bot. Nestes casos
    # conferimos APENAS os 2 primeiros dígitos. UCs normais mantêm prefixo + sufixo.
    uc_curta = len(uc_arq) <= 5
    m = re.search(r"(\d+)\*+(\d+)", texto_item)
    if m:
        prefixo_uc = m.group(1)
        sufixo_uc = m.group(2)
        bate_prefixo = prefixo_uc.startswith(
            primeiros_arq) or primeiros_arq.startswith(prefixo_uc)
        bate_sufixo = sufixo_uc.endswith(
            ultimos_arq) or ultimos_arq.endswith(sufixo_uc)
        resultado = bate_prefixo if uc_curta else (bate_prefixo and bate_sufixo)
        print(f"    [uc_bate] {uc_arq} vs '{m.group(0)}' "
              f"(curta={uc_curta}) → {resultado}")
        return resultado
    blocos = [b for b in re.findall(r"\d+", texto_item) if len(b) >= 2]
    if not blocos:
        print(f"    [uc_bate] {uc_arq}: sem máscara nem blocos → False")
        return False
    achou_primeiro = blocos[0].startswith(primeiros_arq)
    achou_ultimo = any(b.endswith(ultimos_arq) for b in blocos)
    resultado = achou_primeiro if uc_curta else (achou_primeiro and achou_ultimo)
    print(f"    [uc_bate] {uc_arq} vs blocos {blocos} "
          f"(curta={uc_curta}) → {resultado}")
    return resultado


def _eh_avaliacao(texto):
    """
    True quando o bot está exibindo a tela de avaliação do atendimento
    ("Para melhorar nosso serviço..." ou pedido de nota). Isso indica que o
    atendimento foi ENCERRADO — não é um erro. O chamador decide o resultado
    real (Ativa / Ativa e Fatura baixada / Inativa) conforme o que ocorreu antes.
    """
    t = (texto or "").lower()
    return ("melhorar nosso serviço" in t or "melhorar nosso servico" in t or
            "nota válida" in t or "nota valida" in t or "nota inválida" in t or
            "o que achou deste atendimento" in t or
            "o que achou desse atendimento" in t)


def detectar_erro_bot(texto):
    t = texto.lower()
    if "nenhuma fatura vencida" in t or "fatura vencida ou em aberto" in t:
        return True, "sem fatura em aberto"
    if ("dados estão desatualizados" in t or "dados desatualizados" in t or
            "atualizacaowhatsapp" in t or "atualização cadastral" in t or
            "atualizacao cadastral" in t or
            "não consegui localizar o cadastro" in t or
            "nao consegui localizar o cadastro" in t):
        return True, "inativa - dados desatualizados"
    if (FRASE_FALTOU_NUMERO in t or FRASE_FALTOU_NUMERO_ALT in t or
            ("faltou" in t and ("cpf" in t or "cnpj" in t or "número" in t or "numero" in t))):
        return True, "Erro CPF - Verificar zero à esquerda"
    if "melhorar nosso serviço" in t or "melhorar nosso servico" in t:
        return True, "bot pediu avaliação após encerramento"
    if "nota válida" in t or "nota valida" in t or "nota inválida" in t:
        return True, "bot em modo avaliação (nota inválida)"
    if "não entendi o que você disse" in t and ("ver opções" in t or "ver opcoes" in t):
        return True, "bot travado em menu (não entendeu mensagem)"
    if "não entendi o que você disse" in t or "nao entendi" in t:
        return True, "bot não entendeu a mensagem"
    if "não estamos conseguindo nos entender" in t or "nao estamos conseguindo" in t:
        return True, "bot reiniciou conversa (não estamos nos entendendo)"
    if "outros canais de atendimento" in t or "teleatendimento" in t:
        return True, "bot redirecionou para outros canais"
    return False, ""


def verificar_menu_servicos(texto):
    if not texto:
        return False
    t = texto.lower()
    return FRASE_MENU_SERVICOS in t and SAUDACAO_NEOENERGIA not in t


def verificar_pergunta_dividas(texto):
    if not texto:
        return False
    t = texto.lower()
    return FRASE_DIVIDAS in t or FRASE_DIVIDAS_ALT in t


def verificar_mensagem_pix(texto):
    if not texto:
        return False
    return FRASE_PIX in texto.lower()


def verificar_msg_info_pagamento(texto):
    if not texto:
        return False
    t = texto.lower()
    return FRASE_INFO_PAGAMENTO in t or FRASE_INFO_PAGAMENTO_ALT in t


def verificar_msg_multiplas_faturas(texto):
    if not texto:
        return False
    t = texto.lower()
    return FRASE_ESCOLHA_FATURA in t or FRASE_ESCOLHA_FATURA_ALT in t


def verificar_encerramento(texto):
    if not texto:
        return False
    t = texto.lower()
    return any(p in t for p in FRASES_ENCERRAMENTO)


# ──────────────────────────────────────────────────────────────────────────────
# PLAYWRIGHT / WHATSAPP
# ──────────────────────────────────────────────────────────────────────────────

def _ler_msgs_js(page):
    """
    Lê mensagens recebidas diretamente do DOM via JavaScript.
    Retorna dict: {count: int, ultimo: str, todos: list[str], debug: str}
    Funciona mesmo quando as classes CSS são ofuscadas pelo WhatsApp.
    """
    try:
        resultado = page.evaluate(_JS_LER_MSGS)
        if isinstance(resultado, dict):
            # Log enxuto: só imprime quando o estado muda de verdade (nova
            # mensagem ou contagem diferente). Evita repetir a mesma linha
            # dezenas de vezes durante o polling de meio em meio segundo.
            dbg = resultado.get("debug", "")
            cnt = resultado.get("count", 0)
            ult = str(resultado.get("ultimo", ""))[:60]
            assinatura = (dbg, cnt, ult)
            if dbg and assinatura != getattr(_ler_msgs_js, "_ultima_assinatura", None):
                _ler_msgs_js._ultima_assinatura = assinatura
                print(f"  [JS] {dbg} | count={cnt} | ultimo={ult!r}")
            return resultado
    except Exception as e:
        print(f"  [JS] Erro ao ler mensagens: {e}")
    return {"count": 0, "ultimo": "", "todos": [], "debug": "exception"}


# Assinatura da última leitura impressa (para deduplicar logs de polling)
_ler_msgs_js._ultima_assinatura = None


def _all_msg_recebidas(page):
    """
    Retorna lista de elementos Playwright de mensagens recebidas.
    Usada por funções que precisam interagir com o DOM (clicar, scrollar).
    Tenta múltiplos seletores em cascata.
    """
    candidatos = [
        "#main div.message-in",
        "#main div[class*='message-in']",
        "#main [role='row']",
    ]
    for sel in candidatos:
        try:
            loc = page.locator(sel)
            if loc.count() == 0:
                continue
            itens = loc.all()
            if "role='row'" in sel or 'role="row"' in sel:
                # Filtra linhas de saída pelos ícones de entrega
                filtrados = []
                for item in itens:
                    try:
                        tem_check = item.locator(
                            '[data-icon="msg-check"],[data-icon="msg-dblcheck"],'
                            '[data-icon="msg-dblcheck-ack"],[data-icon="msg-time"]'
                        ).count() > 0
                        tem_texto = item.locator(
                            '[data-testid="selectable-text"], span.copyable-text'
                        ).count() > 0
                        if not tem_check and tem_texto:
                            filtrados.append(item)
                    except Exception:
                        continue
                if filtrados:
                    return filtrados
            else:
                return itens
        except Exception:
            continue
    return []


def _ultima_msg_elem(page):
    """Retorna o último elemento Playwright de mensagem recebida, ou None."""
    msgs = _all_msg_recebidas(page)
    return msgs[-1] if msgs else None


def aguardar_whatsapp_carregar(page):
    print("Aguardando WhatsApp Web carregar...")
    try:
        page.wait_for_load_state("networkidle", timeout=60_000)
        print("  network idle atingido.")
    except Exception:
        pass
    SINAIS_PRONTO = [
        'input[role="textbox"][data-tab="3"]',
        'input[aria-label="Pesquisar ou começar uma nova conversa"]',
        'input[aria-label="Search or start new chat"]',
        'div[contenteditable="true"][data-tab="3"]',
        'div[contenteditable="true"]',
        'div[data-testid="chat-list"]',
        'div[data-testid="default-user"]',
        'span[data-testid="default-user"]',
    ]
    limite = time.time() + 300
    qr_avisado = False
    tentativa = 0
    while time.time() < limite:
        tentativa += 1
        for sinal in SINAIS_PRONTO:
            try:
                if page.locator(sinal).count() > 0:
                    print(f"WhatsApp Web pronto (sinal: {sinal}).")
                    time.sleep(1)
                    return True
            except Exception:
                continue
        try:
            tem_canvas = page.locator("canvas").count() > 0
            tem_qr = page.locator('[data-testid="qrcode"]').count() > 0
            if (tem_canvas or tem_qr) and not qr_avisado:
                print("QR Code detectado — escaneie com o celular para continuar.")
                qr_avisado = True
        except Exception:
            pass
        if tentativa % 30 == 0:
            restante = int(limite - time.time())
            print(
                f"  Ainda aguardando WhatsApp Web... ({restante}s restantes)")
        time.sleep(1)
    print("ERRO: WhatsApp Web não carregou dentro de 5 minutos.")
    return False


def caixa_msg(page, _tentou_reabrir=False):
    for sel in SELETORES_CAIXA_MSG:
        try:
            loc = page.locator(sel)
            if loc.count() > 0:
                return loc.last
        except Exception:
            continue
    # Não achou a caixa de mensagem. Causa comum: a conversa fechou/perdeu o foco
    # (após um painel modal, ou navegação). Tenta reabrir UMA vez antes de desistir;
    # isso evita a cascata de falhas em todas as UCs seguintes.
    if not _tentou_reabrir:
        print("  [caixa_msg] Caixa não encontrada — tentando reabrir a conversa...")
        try:
            if _reabrir_conversa_se_necessario(page):
                return caixa_msg(page, _tentou_reabrir=True)
        except Exception as e:
            print(f"  [caixa_msg] Reabertura falhou: {e}")
    # Inclui na mensagem QUAIS seletores foram buscados (alimenta a coluna 'Busca').
    seletores_str = " | ".join(SELETORES_CAIXA_MSG)
    raise RuntimeError(
        f"Não encontrei a caixa de mensagem. Seletores buscados: {seletores_str}")


def enviar(page, texto):
    box = caixa_msg(page)
    box.click()
    time.sleep(0.2)
    box.type(str(texto), delay=15)
    page.keyboard.press("Enter")
    time.sleep(0.8)


def qtd_msgs(page):
    try:
        return _ler_msgs_js(page)["count"]
    except Exception:
        return 0


def ultima_msg_texto(page):
    try:
        return _ler_msgs_js(page)["ultimo"]
    except Exception:
        return ""


def aguardar_bot(page, qtd_antes, texto_antes, timeout=TIMEOUT_BOT, palavras_chave=None):
    palavras_chave = [p.lower() for p in (palavras_chave or [])]
    limite = time.time() + timeout
    texto_antes_lower = texto_antes.lower() if texto_antes else ""

    while time.time() < limite:
        try:
            dados = _ler_msgs_js(page)
        except Exception:
            time.sleep(0.5)
            continue

        qtd_atual = dados.get("count", 0)
        ultimo_atual = dados.get("ultimo", "")
        ultimo_lower = ultimo_atual.lower()
        todos_recentes = dados.get("todos", [])

        # Saída imediata: bot pediu avaliação
        if "melhorar nosso serviço" in ultimo_lower or "melhorar nosso servico" in ultimo_lower:
            return ultimo_atual

        mudou = (qtd_atual > qtd_antes) or (
            ultimo_lower and ultimo_lower != texto_antes_lower)

        if mudou:
            # Espera o bot terminar de digitar (bot pode enviar várias mensagens em sequência)
            time.sleep(3.0)
            dados2 = _ler_msgs_js(page)
            resposta = dados2.get("ultimo", ultimo_atual)
            todos2 = dados2.get("todos", todos_recentes)

            # Atualiza estado → evita reler a mesma mensagem em loop
            qtd_antes = dados2.get("count", qtd_atual)
            texto_antes_lower = resposta.lower()

            if "melhorar nosso serviço" in resposta.lower() or "melhorar nosso servico" in resposta.lower():
                return resposta

            if not palavras_chave:
                return resposta

            # Verifica em TODAS as mensagens recentes — não só na última.
            # Isso resolve o caso em que a keyword está na penúltima mensagem
            # e a última é "Novidade! basta nos mandar um áudio" (sem keyword).
            texto_concat = "\n".join(todos2).lower()
            for p in palavras_chave:
                if p in texto_concat:
                    return resposta

            # Keyword não encontrada ainda — continua esperando.
            # Estado já atualizado, próxima mudança será detectada corretamente.

        time.sleep(0.5)
    return ""


def _digitar_na_busca(page, busca_loc, texto):
    busca_loc.click()
    time.sleep(0.3)
    try:
        busca_loc.fill("")
    except Exception:
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
    try:
        busca_loc.fill(texto)
    except Exception:
        busca_loc.type(texto, delay=30)
    time.sleep(2.5)


def _localizar_caixa_busca(page):
    """
    Localiza a caixa de busca/pesquisa do WhatsApp tentando todos os seletores.
    Retorna (locator, seletor_usado) ou (None, mensagem_do_que_foi_buscado).
    Mesma filosofia da caixa_msg: quando não encontra, devolve a lista de
    seletores tentados para alimentar diagnósticos / coluna 'Busca'.
    """
    for sel in SELETORES_CAIXA_BUSCA:
        try:
            loc = page.locator(sel)
            if loc.count() > 0:
                return loc.first, sel
        except Exception:
            continue
    seletores_str = " | ".join(SELETORES_CAIXA_BUSCA)
    return None, f"Não encontrei a caixa de busca. Seletores buscados: {seletores_str}"


def abrir_conversa(page, nome_contato):
    print(f"Abrindo conversa: {nome_contato}")
    busca, sel_usado = _localizar_caixa_busca(page)
    if busca is None:
        # sel_usado contém a mensagem com os seletores tentados
        print(f"ERRO: {sel_usado}")
        return False
    print(f"  Busca encontrada: {sel_usado}")
    try:
        _digitar_na_busca(page, busca, nome_contato)
    except Exception as e:
        print(f"ERRO ao digitar na busca: {e}")
        return False
    try:
        page.locator(f'span[title="{nome_contato}"]').first.click(
            timeout=10_000)
        time.sleep(2)
        print(f"  Conversa '{nome_contato}' aberta.")
        return True
    except Exception:
        pass
    try:
        page.keyboard.press("Enter")
        time.sleep(3)
        print("  Conversa aberta via ENTER.")
        return True
    except Exception as e:
        print(f"ERRO fallback abrir conversa: {e}")
        return False


def enviar_sair_e_aguardar_encerramento(page, max_tentativas=2):
    for tentativa in range(1, max_tentativas + 1):
        try:
            dados_antes = _ler_msgs_js(page)
            n = dados_antes.get("count", 0)
            t = dados_antes.get("ultimo", "").lower()
            print(
                f"  [Sair] Tentativa {tentativa}/{max_tentativas} | msgs antes={n} | ultimo={t[:50]!r}")
            enviar(page, "Sair")
            timeout = TIMEOUT_BOT if tentativa == 1 else TIMEOUT_REENVIO
            r = aguardar_bot(page, qtd_antes=n, texto_antes=t,
                             timeout=timeout, palavras_chave=FRASES_ENCERRAMENTO)
            # Verifica encerramento em r E em todas as mensagens recentes
            dados_depois = _ler_msgs_js(page)
            todos_recentes = dados_depois.get("todos", [])
            print(f"  [Sair] Resposta aguardar_bot: {r[:80]!r}")
            print(
                f"  [Sair] Todos recentes ({len(todos_recentes)}): {[t[:50] for t in todos_recentes]}")
            texto_concat = "\n".join(todos_recentes).lower()
            if verificar_encerramento(r) or any(p in texto_concat for p in FRASES_ENCERRAMENTO):
                print("  [Sair] Encerramento confirmado.")
                return True
            print(
                f"  [Sair] Encerramento não detectado. Tentativa {tentativa}/{max_tentativas}.")
        except Exception as e:
            print(f"  [Sair] erro: {e}")
    return False


def contar_pdfs(page):
    NEG_PIX = [
        "copiar código pix", "copiar codigo pix",
        "para pagar com pix", "br.gov.bcb.pix",
        "00020126", "pix copia e cola",
    ]
    POS_PDF = [
        "segunda via -", "fatura_elektro", ".pdf", "pdf ·",
    ]
    seletores_pdf = [
        '#main [role="row"]:has-text(".pdf")',
        '#main [role="row"]:has-text("Segunda via -")',
        '#main [role="row"]:has-text("fatura_Elektro")',
        '#main [role="row"]:has-text("fatura_elektro")',
        '#main [role="row"]:has-text("PDF ·")',
        '#main div.message-in:has-text(".pdf")',
        '#main div.message-in:has-text("Segunda via -")',
        '#main div[class*="message-in"]:has-text(".pdf")',
    ]
    max_pdfs = 0
    for seletor in seletores_pdf:
        try:
            msgs = page.locator(seletor).all()
            validas = 0
            for m in msgs:
                try:
                    txt = m.inner_text().lower()
                    eh_pix = any(neg in txt for neg in NEG_PIX)
                    eh_pdf = any(pos in txt for pos in POS_PDF)
                    if eh_pix and not eh_pdf:
                        continue
                    validas += 1
                except Exception:
                    validas += 1
            max_pdfs = max(max_pdfs, validas)
        except Exception:
            pass
    if max_pdfs == 0:
        try:
            html_main = page.evaluate(
                "() => { const main = document.querySelector('#main'); return main ? main.innerHTML : ''; }")
            html_lower = (html_main or "").lower()
            ocorrencias_segunda_via = html_lower.count("segunda via -")
            ocorrencias_fatura_elektro = html_lower.count("fatura_elektro")
            max_pdfs = max(ocorrencias_segunda_via, ocorrencias_fatura_elektro)
        except Exception:
            pass
    return max_pdfs


def existe_pdf_recente(page, pdfs_antes=0):
    return contar_pdfs(page) > pdfs_antes


def _pdf_no_dom(page):
    seletores_dom = [
        '#main [role="row"]:has(span[data-icon="document"])',
        '#main [role="row"]:has([data-testid="document-thumb"])',
        '#main [role="row"]:has(div[class*="document-body"])',
        f'{SEL_MSG_RECEBIDA}:has(span[data-icon="document"])',
        f'{SEL_MSG_RECEBIDA}:has([data-testid="document-thumb"])',
    ]
    for sel in seletores_dom:
        try:
            if page.locator(sel).count() > 0:
                return True
        except Exception:
            pass
    try:
        resultado = page.evaluate("""
            () => {
                const main = document.querySelector('#main');
                if (!main) return false;
                // Procura em qualquer mensagem (entrada ou saída) por PDF — depois filtra PIX
                const candidatos = Array.from(main.querySelectorAll(
                    'div.message-in, div[class*="message-in"], [role="row"]'
                ));
                const vistos = new Set();
                for (const msg of candidatos) {
                    if (vistos.has(msg)) continue;
                    vistos.add(msg);
                    const html = msg.innerHTML.toLowerCase();
                    if (html.includes('br.gov.bcb.pix') || html.includes('00020126') || html.includes('pix copia e cola')) continue;
                    if (html.includes('segunda via -') || html.includes('fatura_elektro') ||
                        (html.includes('.pdf') && !html.includes('copycode'))) {
                        return true;
                    }
                    if (msg.querySelector('[data-testid="document-thumb"], span[data-icon="document"]')) {
                        return true;
                    }
                }
                return false;
            }
        """)
        if resultado:
            return True
    except Exception:
        pass
    return False


def baixar_pdf(page, caminho_destino):
    PASTA_FATURAS.mkdir(parents=True, exist_ok=True)
    print("Tentando baixar o PDF...")
    time.sleep(2)

    def _encontrar_msg_pdf():
        seletores_tem_doc = [
            '#main [role="row"]:has(span[data-icon="document"])',
            '#main [role="row"]:has([data-testid="document-thumb"])',
            '#main [role="row"]:has(div[class*="document"])',
            f'{SEL_MSG_RECEBIDA}:has(span[data-icon="document"])',
            f'{SEL_MSG_RECEBIDA}:has([data-testid="document-thumb"])',
            f'{SEL_MSG_RECEBIDA}:has(div[class*="document"])',
            f'{SEL_MSG_RECEBIDA}:has(div[class*="media-button"])',
        ]
        for sel in seletores_tem_doc:
            try:
                loc = page.locator(sel)
                if loc.count() > 0:
                    return loc.last
            except Exception:
                pass
        for seletor in [
            '#main [role="row"]:has-text("Segunda via -")',
            '#main [role="row"]:has-text("fatura_Elektro")',
            '#main [role="row"]:has-text("fatura_elektro")',
            '#main [role="row"]:has-text(".pdf")',
            'div.message-in:has-text("Segunda via -")',
            'div.message-in:has-text("fatura_elektro")',
            'div.message-in:has-text(".pdf")',
        ]:
            try:
                loc = page.locator(seletor)
                if loc.count() > 0:
                    return loc.last
            except Exception:
                pass
        try:
            msgs = _all_msg_recebidas(page)
            for msg_item in reversed(msgs):
                for sel_doc in [
                    'span[data-icon="document"]',
                    '[data-testid="document-thumb"]',
                    'div[class*="document"]',
                    'div[class*="media-button"]',
                ]:
                    try:
                        if msg_item.locator(sel_doc).count() > 0:
                            return msg_item
                    except Exception:
                        continue
        except Exception:
            pass
        return None

    msg = _encontrar_msg_pdf()
    if msg is None:
        print("Não encontrei mensagem com PDF.")
        return False

    msg.scroll_into_view_if_needed()
    time.sleep(0.5)
    print("Estratégia 1: clicando no card do PDF...")
    _seletores_card_pdf = [
        'span[data-icon="document"]',
        '[data-testid="document-thumb"]',
        'div[class*="document"]',
        'div[class*="media-button"]',
    ]
    clicou_card = False
    for sel in _seletores_card_pdf:
        try:
            card = msg.locator(sel)
            if card.count() > 0:
                try:
                    with page.expect_download(timeout=6000) as dl_info:
                        card.first.click(timeout=5000)
                    download = dl_info.value
                    download.save_as(str(caminho_destino))
                    print(
                        f"PDF salvo diretamente (clique no card): {caminho_destino}")
                    return True
                except Exception:
                    clicou_card = True
                    print(f"  Card clicado com seletor: {sel} (viewer)")
                    break
        except Exception:
            continue

    if not clicou_card:
        try:
            try:
                with page.expect_download(timeout=6000) as dl_info:
                    msg.locator('div[role="button"]').first.click(timeout=5000)
                download = dl_info.value
                download.save_as(str(caminho_destino))
                print(
                    f"PDF salvo diretamente (div[role=button]): {caminho_destino}")
                return True
            except Exception:
                msg.locator('div[role="button"]').first.click(timeout=5000)
                clicou_card = True
                print("  Card clicado via div[role=button]")
        except Exception:
            pass

    if clicou_card:
        time.sleep(2)
        _seletores_download_viewer = [
            'span[data-icon="download-alt"]',
            'span[data-icon="download"]',
            '[aria-label="Baixar"]',
            '[aria-label="Download"]',
            '[title="Baixar"]',
            '[title="Download"]',
            '[data-testid="download-button"]',
            'button[aria-label*="ownload"]',
            'div[aria-label*="ownload"]',
        ]
        for sel in _seletores_download_viewer:
            try:
                btn = page.locator(sel)
                if btn.count() > 0:
                    print(f"  Botão de download encontrado: {sel}")
                    with page.expect_download(timeout=25000) as dl_info:
                        btn.last.click(timeout=5000)
                    download = dl_info.value
                    download.save_as(str(caminho_destino))
                    print(f"PDF salvo em: {caminho_destino}")
                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.5)
                    except Exception:
                        pass
                    return True
            except Exception as e:
                print(f"  {sel} falhou: {e}")
                continue
        try:
            page.keyboard.press("Escape")
            time.sleep(0.5)
        except Exception:
            pass

    print("Estratégia 2: menu de contexto da mensagem...")
    try:
        msg.hover(timeout=5000)
    except Exception as e:
        print(f"  hover falhou/timeout ({e}) — seguindo mesmo assim.")
    time.sleep(0.8)
    _seletores_menu = [
        'span[data-icon="down-context"]',
        'span[data-icon="chevron-down-outline"]',
        'span[data-icon="menu"]',
    ]
    clicou_menu = False
    for sel in _seletores_menu:
        try:
            botao = msg.locator(sel)
            if botao.count() > 0:
                botao.last.click(timeout=5000)
                clicou_menu = True
                print(f"  Menu aberto com: {sel}")
                break
        except Exception:
            continue
    if clicou_menu:
        time.sleep(0.6)
        try:
            with page.expect_download(timeout=25000) as dl_info:
                try:
                    page.get_by_text(
                        "Baixar", exact=True).last.click(timeout=5000)
                except Exception:
                    page.get_by_text(
                        "Download", exact=True).last.click(timeout=5000)
            download = dl_info.value
            download.save_as(str(caminho_destino))
            print(f"PDF salvo em: {caminho_destino}")
            return True
        except Exception as e:
            print(f"  Menu/Baixar falhou: {e}")
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass

    print("Estratégia 3: procurando link blob: diretamente no DOM...")
    try:
        blob_url = page.evaluate("""
            () => {
                const links = Array.from(document.querySelectorAll('a[href^="blob:"]'));
                return links.length > 0 ? links[links.length - 1].href : null;
            }
        """)
        if blob_url:
            print(f"  blob URL encontrada: {str(blob_url)[:60]}...")
            with page.expect_download(timeout=25000) as dl_info:
                page.evaluate(f"""
                    () => {{
                        const a = document.createElement('a');
                        a.href = '{blob_url}';
                        a.download = 'fatura_elektro.pdf';
                        document.body.appendChild(a);
                        a.click();
                        document.body.removeChild(a);
                    }}
                """)
            download = dl_info.value
            download.save_as(str(caminho_destino))
            print(f"PDF salvo em: {caminho_destino}")
            return True
    except Exception as e:
        print(f"  Estratégia blob URL falhou: {e}")

    print("Todas as estratégias de download falharam.")
    return False


# ──────────────────────────────────────────────────────────────────────────────
# INTERAÇÕES COM O BOT
# ──────────────────────────────────────────────────────────────────────────────

def tratar_pergunta_dividas(page, resposta):
    if not verificar_pergunta_dividas(resposta):
        return resposta
    print("  [Dívidas] Bot perguntou sobre dívidas — respondendo 'Isso mesmo'...")
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    enviar(page, "Isso mesmo")
    nova = aguardar_bot(page, qtd_antes=n, texto_antes=t, timeout=TIMEOUT_BOT,
                        palavras_chave=["fatura", "pdf", ".pdf", "unidade", "opção",
                                        "opcao", "escolha", "selecione", "pix", "⚡",
                                        "encerrado", "sair", "nenhuma fatura"])
    print(f"  [Dívidas] Bot após 'Isso mesmo': {nova[:120]}")
    return nova


def _saudacao_visivel_na_tela(page):
    try:
        ultima = ultima_msg_texto(page).lower()
        if SAUDACAO_NEOENERGIA in ultima:
            return True
    except Exception:
        pass
    return False


def _limpar_estado_residual(page, max_ciclos=3):
    """
    Garante que NÃO estamos presos no estado do atendimento anterior antes de
    iniciar um novo. Se a tela mostra a tela de avaliação ("melhorar nosso
    serviço") ou uma mensagem de encerramento, envia 'Sair' para destravar o bot.
    Retorna o texto da última mensagem após a limpeza.
    """
    for ciclo in range(1, max_ciclos + 1):
        ultima = ultima_msg_texto(page)
        ul = ultima.lower()
        em_avaliacao = _eh_avaliacao(ultima)
        em_encerramento = verificar_encerramento(ultima)
        if not em_avaliacao and not em_encerramento:
            return ultima  # estado já limpo (ou esperando entrada do usuário)
        if em_avaliacao:
            print(f"  [P1-limpeza {ciclo}] Tela em avaliação — enviando 'Sair' "
                  f"para destravar o atendimento anterior...")
        else:
            print(f"  [P1-limpeza {ciclo}] Tela em encerramento — enviando 'Sair' "
                  f"para garantir estado limpo...")
        try:
            enviar_sair_e_aguardar_encerramento(page)
        except Exception as e:
            print(f"  [P1-limpeza] erro ao enviar 'Sair': {e}")
        time.sleep(2)
    return ultima_msg_texto(page)


def validar_saudacao_neoenergia(page, max_tentativas=4):
    _PALAVRAS_SAUDACAO = [
        "assistente virtual", "neoenergia elektro", "escolha o serviço",
        "neoenergia", "elektro",
    ]
    # ── Passo 0: limpar estado residual do atendimento anterior ──────────────
    # Sem isso, a tela de avaliação/encerramento do atendimento anterior fica na
    # tela e induzimos um falso "saudação encontrada" usando a saudação ANTIGA do
    # histórico — o bug visto no run (P1 dá OK, P2 cai em avaliação → erro).
    _limpar_estado_residual(page)

    # Se, após a limpeza, a saudação Neoenergia está REALMENTE na tela (última
    # mensagem), aceitamos sem reenviar.
    if _saudacao_visivel_na_tela(page):
        print("  [P1] Saudação Neoenergia já está na tela ✓ (sem reenviar)")
        return True, ""

    for tentativa in range(1, max_tentativas + 1):
        print(
            f"[P1] Tentativa {tentativa}/{max_tentativas} — enviando 'Olá'...")
        n = qtd_msgs(page)
        t = ultima_msg_texto(page)
        enviar(page, "Olá")
        r = aguardar_bot(page, n, t, timeout=TIMEOUT_BOT,
                         palavras_chave=_PALAVRAS_SAUDACAO + [
                             "olá", "oi", "bem-vindo", "cpf", "menu",
                             "não entendi", "nao entendi", "avaliar", "nota válida",
                             "conte sempre com a gente", "estou por aqui",
                             "dica de segurança", "dica de seguranca",
                         ])
        print(f"  Bot: {r[:160]}")

        # Aceita a saudação SOMENTE se ela for a resposta nova ao nosso 'Olá'
        # (em r) ou estiver entre as 2 últimas mensagens da tela. NÃO varremos
        # todo o histórico — assim não confundimos com a saudação do atendimento
        # anterior, que continuaria lá no meio das mensagens antigas.
        rl = r.lower()
        if SAUDACAO_NEOENERGIA in rl:
            print("  [P1] Saudação Neoenergia recebida ✓")
            return True, ""
        ultimas_2 = _ler_msgs_js(page).get("todos", [])[-2:]
        if any(SAUDACAO_NEOENERGIA in t.lower() for t in ultimas_2):
            print("  [P1] Saudação Neoenergia entre as 2 últimas mensagens ✓")
            return True, ""

        if not r:
            print("  Sem resposta do bot — tentando reiniciar com 'Sair'...")
            try:
                enviar_sair_e_aguardar_encerramento(page)
            except Exception:
                pass
            time.sleep(2)
            continue
        # Se a resposta ao 'Olá' foi avaliação/encerramento, o bot ainda estava
        # preso no estado anterior — limpa e tenta de novo.
        if _eh_avaliacao(r) or verificar_encerramento(r):
            print("  [P1] Resposta ainda é avaliação/encerramento — limpando e repetindo...")
            _limpar_estado_residual(page)
            time.sleep(2)
            continue
        if verificar_encerramento(r):
            print(
                "  [P1] Bot enviou encerramento (estado limpo) — reenviando 'Olá'...")
            time.sleep(2)
            continue
        eh_erro, _ = detectar_erro_bot(r)
        if eh_erro or "não entendi" in rl or "nao entendi" in rl:
            print(f"  [P1] Erro detectado — enviando 'Sair' para reiniciar...")
            try:
                enviar_sair_e_aguardar_encerramento(page)
            except Exception:
                pass
            time.sleep(2)
            continue

        # Resposta inesperada: antes de resetar, aguarda mais 5s para ver
        # se o bot ainda está enviando mensagens em sequência
        print(
            f"  [P1] Resposta não reconhecida: '{r[:80]}' — aguardando 5s antes de resetar...")
        time.sleep(5)
        todos_apos_espera = _ler_msgs_js(page).get("todos", [])
        if any(SAUDACAO_NEOENERGIA in t.lower() for t in todos_apos_espera):
            print("  [P1] Saudação chegou durante a espera extra ✓")
            return True, ""
        print("  [P1] Saudação não chegou — reiniciando com 'Sair'...")
        try:
            enviar_sair_e_aguardar_encerramento(page)
        except Exception:
            pass
        time.sleep(2)
    return False, "Erro: saudação Neoenergia não recebida"


def _localizar_bolha_uc(page):
    try:
        msgs = _all_msg_recebidas(page)
        for msg in reversed(msgs):
            try:
                if "unidade consumidora" in msg.inner_text().lower():
                    return msg
            except Exception:
                continue
    except Exception:
        pass
    return None


def _clicar_sim_na_bolha(page, bolha_loc, n_antes, t_antes):
    clicou = False
    if bolha_loc:
        try:
            botoes = bolha_loc.locator("button").all()
            for btn in botoes:
                try:
                    txt = btn.inner_text().strip().lower()
                    if txt in ("sim", "yes", "s"):
                        btn.click(timeout=6_000)
                        clicou = True
                        print("  Botão 'Sim' clicado na bolha.")
                        break
                except Exception:
                    continue
        except Exception:
            pass
    if not clicou:
        print("  Botão 'Sim' não encontrado — digitando no chat.")
        enviar(page, "Sim")
    return aguardar_bot(page, n_antes, t_antes, timeout=TIMEOUT_BOT,
                        palavras_chave=["1", "fatura", "opção", "opcao", "escolha", "menu",
                                        "sair", "segunda via", "⚡", "dívida", "divida",
                                        "consultar", "pix"])


def processar_uc_unica(page, uc_arquivo, texto_uc=None):
    """
    texto_uc: texto já capturado pelo aguardar_bot do P4 — evita re-consulta ao DOM
    que pode falhar com elementos stale após re-renderização do WhatsApp.
    """
    bolha = _localizar_bolha_uc(page)  # ainda usado para clicar no botão Sim

    # Ordem de preferência para obter o texto da UC:
    # 1) texto passado diretamente de P4 (mais confiável)
    # 2) bolha.inner_text() com fallback robusto
    # 3) ultima_msg_texto() via JS
    if texto_uc:
        texto_bolha = texto_uc
    else:
        texto_bolha = ""
        if bolha:
            try:
                texto_bolha = bolha.inner_text().strip()
            except Exception as e:
                print(f"  [UC única] bolha.inner_text() falhou: {e} — usando JS")
        if not texto_bolha:
            dados_js = _ler_msgs_js(page)
            todos = dados_js.get("todos", [])
            texto_bolha = next(
                (t for t in reversed(todos) if "unidade consumidora" in t.lower()),
                dados_js.get("ultimo", "")
            )

    print(f"  [UC única] Texto da bolha: {texto_bolha[:160]}")
    if not texto_bolha:
        return False, "ERRO: não foi possível ler o texto da UC"
    if not uc_bate(uc_arquivo, texto_bolha):
        print(f"  UC arquivo ({uc_arquivo}) NÃO bate.")
        return False, "UC não confere com a UC apresentada pelo bot"
    print("  UC compatível. Confirmando...")
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    r = _clicar_sim_na_bolha(page, bolha, n, t)
    print(f"  Resposta após Sim: {r[:120]}")
    if not r:
        return False, "ERRO: bot não confirmou UC após 'Sim'"
    return True, r


def _extrair_linha_uc(texto_item):
    if not texto_item:
        return ""
    for linha in texto_item.split("\n"):
        linha = linha.strip()
        if not linha:
            continue
        if re.match(r"(?i)^unidade\s+\d+", linha):
            return linha
        if re.search(r"\d+\*+\d+", linha):
            return linha
    for linha in texto_item.split("\n"):
        if linha.strip():
            return linha.strip()
    return texto_item.strip()


def _listar_itens_painel(page):
    seletores_itens = [
        '[role="listitem"]', '[role="option"]', '[role="radio"]',
        '[role="gridcell"]',
        'div[data-testid="cell-frame-container"]',
        'div[data-testid="cell-frame-primary"]',
    ]
    for sel in seletores_itens:
        try:
            todos = page.locator(sel).all()
            candidatos = []
            for it in todos:
                try:
                    txt = it.inner_text()
                    txt_lower = txt.lower()
                    if (re.search(r"\d+\*+\d+", txt) or
                            re.match(r"(?i).*unidade\s+\d+", txt_lower) or
                            re.match(r"(?i).*fatura\s+\d+", txt_lower)):
                        candidatos.append(it)
                except Exception:
                    continue
            if candidatos:
                print(
                    f"  [painel] {len(candidatos)} item(ns) encontrado(s) com seletor: {sel}")
                return candidatos
        except Exception:
            continue
    print("  [painel] Nenhum item encontrado.")
    return []


def _clicar_no_item_uc(page, item_alvo):
    estrategias = [
        ('input[type="radio"]', 'input radio'),
        ('[role="radio"]', 'role=radio'),
        ('div[class*="radio"]', 'div.radio'),
        ('span[class*="radio"]', 'span.radio'),
        (None, 'container do item'),
    ]
    for seletor, descricao in estrategias:
        try:
            if seletor is None:
                alvo = item_alvo
            else:
                radio = item_alvo.locator(seletor).first
                if radio.count() == 0:
                    continue
                alvo = radio
            alvo.scroll_into_view_if_needed(timeout=3000)
            time.sleep(0.3)
            alvo.click(timeout=4000)
            print(f"  [painel] ✓ Clique disparado em '{descricao}'")
            time.sleep(0.8)
            return True
        except Exception as e:
            print(f"  [painel] '{descricao}' falhou: {e}")
            continue
    return False


def _confirmar_selecao_painel(page):
    print("  [painel] Procurando botão por data-testid='list-msg-modal-button'...")
    try:
        btn = page.locator('[data-testid="list-msg-modal-button"]')
        count = btn.count()
        print(f"  [painel] {count} elemento(s) com esse testid")
        if count > 0:
            alvo = btn.last
            try:
                alvo.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            time.sleep(0.3)
            alvo.click(timeout=4000)
            print(
                f"  [painel] ✓ Botão clicado via data-testid='list-msg-modal-button'")
            time.sleep(1.5)
            return True
    except Exception as e:
        print(f"  [painel] testid list-msg-modal-button falhou: {e}")
    seletores_confirm = [
        'div[role="button"][data-testid*="modal-button"]',
        'div[role="button"][data-testid*="list-msg"]',
        'span[data-icon="send-light"]', 'span[data-icon="send"]',
        'span[data-icon="send-filled"]',
        'button[aria-label="Enviar"]', 'button[aria-label="Send"]',
        'div[aria-label="Enviar"]', 'div[aria-label="Send"]',
        '[aria-label="Confirmar"]', '[aria-label="Confirm"]',
        'div[data-animate-modal-popup="true"] div[role="button"]',
        'div[data-animate-modal-body="true"] div[role="button"]',
    ]
    for sel in seletores_confirm:
        try:
            btn = page.locator(sel)
            if btn.count() == 0:
                continue
            alvo = btn.last
            try:
                alvo.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            time.sleep(0.3)
            alvo.click(timeout=3000)
            print(f"  [painel] ✓ Botão de confirmação clicado ({sel})")
            time.sleep(1.5)
            return True
        except Exception:
            continue
    try:
        clicked = page.evaluate("""
            () => {
                const verdesWA = ['rgb(37, 211, 102)','rgb(0, 168, 132)','rgb(18, 140, 126)','rgb(7, 94, 84)'];
                const candidatos = Array.from(document.querySelectorAll('button, div[role="button"], [role="button"]'));
                let melhor = null; let melhorArea = 0;
                for (const el of candidatos) {
                    const rect = el.getBoundingClientRect();
                    if (rect.width < 30 || rect.height < 30) continue;
                    const bg = window.getComputedStyle(el).backgroundColor;
                    if (!verdesWA.includes(bg)) continue;
                    const area = rect.width * rect.height;
                    if (area > melhorArea) { melhorArea = area; melhor = el; }
                }
                if (melhor) {
                    melhor.scrollIntoView({block:'center'}); melhor.click();
                    return {success:true, info:`bg=${window.getComputedStyle(melhor).backgroundColor} area=${melhorArea}`};
                }
                return {success:false, info:'nenhum botão verde encontrado'};
            }
        """)
        if clicked and clicked.get("success"):
            print(
                f"  [painel] ✓ Botão verde clicado via JS ({clicked.get('info')})")
            time.sleep(1.5)
            return True
    except Exception as e:
        print(f"  [painel] JS busca verde falhou: {e}")
    try:
        modal = page.locator('div[data-testid="popup-contents"]').last
        if modal.count() == 0:
            modal = page.locator('div[data-animate-modal-popup="true"]').last
        if modal.count() > 0:
            botoes = modal.locator('div[role="button"], button').all()
            if botoes:
                ultimo = botoes[-1]
                try:
                    ultimo.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass
                ultimo.click(timeout=3000)
                print(
                    f"  [painel] ✓ Último botão do modal clicado (de {len(botoes)} botões)")
                time.sleep(1.5)
                return True
    except Exception as e:
        print(f"  [painel] Último botão do modal falhou: {e}")
    print("  [painel] ❌ Botão de confirmação não encontrado em NENHUMA estratégia.")
    return False


def processar_lista_uc(page, uc_arquivo):
    try:
        ultima = ultima_msg_texto(page).lower()
        eh_pre_selecao_uc = (
            "ver unidades" in ultima or
            ("unidades consumidoras" in ultima and
             ("escolha a unidade" in ultima or "selecione a unidade" in ultima or
              "escolha a unidade para" in ultima)) or
            "qual é a sua unidade" in ultima or
            "qual e a sua unidade" in ultima
        )
        if not eh_pre_selecao_uc:
            indicadores_pos_uc = [
                "informações para pagamento", "informacoes para pagamento",
                "para pagar com pix", "ver faturas", "escolha a fatura",
                ".pdf", "fatura digital",
            ]
            if "encontrei" in ultima and "fatura" in ultima:
                indicadores_pos_uc.append("encontrei")
            if any(ind in ultima for ind in indicadores_pos_uc):
                print(
                    f"  [Lista UC] ✓ Estado pós-UC detectado: '{ultima[:80]}...'")
                print(
                    f"  [Lista UC] Pulando abertura do painel — UC já foi confirmada.")
                return True, ultima_msg_texto(page)
    except Exception:
        pass
    print("  [Lista UC] Abrindo painel 'Ver unidades'...")
    clicou_ver = False
    for sel in [
        'button[title="Ver unidades"]',
        'button:has-text("Ver unidades")',
        'div[role="button"]:has-text("Ver unidades")',
    ]:
        try:
            btn_ver = page.locator(sel).last
            if btn_ver.count() > 0:
                btn_ver.click(timeout=6000)
                clicou_ver = True
                print(f"  [Lista UC] Clicou em 'Ver unidades' ({sel})")
                time.sleep(2)
                break
        except Exception:
            continue
    if not clicou_ver:
        return False, "mais de uma UC - botão 'Ver unidades' não encontrado"
    itens = _listar_itens_painel(page)
    if not itens:
        try:
            page.keyboard.press("Escape")
            time.sleep(0.5)
        except Exception:
            pass
        return False, "mais de uma UC - itens não encontrados no painel"
    item_alvo = None
    idx_alvo = None
    for idx, item in enumerate(itens):
        try:
            texto_completo = item.inner_text().strip()
            linha_uc = _extrair_linha_uc(texto_completo)
            print(
                f"    [{idx+1}] linha_uc='{linha_uc}'  (texto completo: {texto_completo[:80]!r})")
            if uc_bate(uc_arquivo, linha_uc):
                item_alvo = item
                idx_alvo = idx + 1
                print(
                    f"    → UC COMPATÍVEL na posição {idx_alvo}: '{linha_uc}'")
                break
        except Exception as e:
            print(f"    Erro ao ler item {idx+1}: {e}")
            continue
    if item_alvo is None:
        print("  [Lista UC] Nenhuma UC do painel bateu com o arquivo.")
        try:
            for sel_f in [
                'button[aria-label="Fechar"]', 'button[aria-label="Close"]',
                'span[data-icon="x"]', 'span[data-icon="close"]',
            ]:
                btn_f = page.locator(sel_f).first
                if btn_f.count() > 0:
                    btn_f.click(timeout=3000)
                    time.sleep(0.6)
                    break
            else:
                page.keyboard.press("Escape")
                time.sleep(0.5)
        except Exception:
            pass
        return False, "mais de uma UC - nenhuma conferiu com o arquivo"
    PALAVRAS_RESP = [
        "posso seguir", "unidade consumidora", "sim", "confirmar",
        "fatura", "opção", "opcao", "menu", "sair",
        "dívida", "divida", "consultar",
        "encontrei", "informações para pagamento",
        "informacoes para pagamento", ".pdf",
    ]
    r = ""
    MAX_TENTATIVAS_CLIQUE = 2
    for tentativa_clique in range(1, MAX_TENTATIVAS_CLIQUE + 1):
        print(
            f"  [Lista UC] Tentativa de seleção {tentativa_clique}/{MAX_TENTATIVAS_CLIQUE}...")
        if tentativa_clique > 1:
            painel_aberto = False
            try:
                modal = page.locator(
                    'div[data-testid="popup-contents"], div[data-animate-modal-popup="true"]')
                painel_aberto = modal.count() > 0
            except Exception:
                painel_aberto = False
            if not painel_aberto:
                print(
                    "  [Lista UC] Reabrindo painel 'Ver unidades' para 2ª tentativa...")
                reabriu = False
                for sel in ['button[title="Ver unidades"]', 'button:has-text("Ver unidades")']:
                    try:
                        btn_ver = page.locator(sel).last
                        if btn_ver.count() > 0 and btn_ver.is_visible():
                            btn_ver.click(timeout=5000)
                            reabriu = True
                            time.sleep(2)
                            break
                    except Exception:
                        continue
                if reabriu:
                    itens2 = _listar_itens_painel(page)
                    item_alvo = None
                    for item in itens2:
                        try:
                            linha_uc = _extrair_linha_uc(
                                item.inner_text().strip())
                            if uc_bate(uc_arquivo, linha_uc):
                                item_alvo = item
                                break
                        except Exception:
                            continue
                    if item_alvo is None:
                        print(
                            "  [Lista UC] Item não reencontrado no painel reaberto.")
                        break
        print(f"  [Lista UC] Clicando no radio da posição {idx_alvo}...")
        if not _clicar_no_item_uc(page, item_alvo):
            print("  [Lista UC] Clique no radio falhou nesta tentativa.")
            time.sleep(1)
            continue
        time.sleep(1)
        n_antes = qtd_msgs(page)
        t_antes = ultima_msg_texto(page)
        print("  [Lista UC] Clicando no botão verde de confirmação...")
        confirmou = _confirmar_selecao_painel(page)
        if not confirmou:
            print("  [Lista UC] Confirmação não encontrada — tentando Escape...")
            try:
                page.keyboard.press("Escape")
                time.sleep(0.8)
            except Exception:
                pass
        r = aguardar_bot(page, n_antes, t_antes,
                         timeout=TIMEOUT_POS_CLIQUE_PAINEL, palavras_chave=PALAVRAS_RESP)
        print(
            f"  [Lista UC] Bot após confirmação: {r[:160] if r else '(sem resposta)'}")
        if r:
            break
        print(f"  [Lista UC] Sem resposta na tentativa {tentativa_clique}.")
    if not r:
        print(
            "  [Lista UC] Painel não respondeu — enviando 'Sair' e marcando para retry.")
        try:
            enviar_sair_e_aguardar_encerramento(page)
        except Exception:
            pass
        return False, "ERRO: painel de UC não respondeu após 2 tentativas de clique"
    if "posso seguir" in r.lower() or "unidade consumidora" in r.lower():
        print(
            "  [Lista UC] Bot pediu confirmação 'Posso seguir?'. Clicando em 'Sim'...")
        bolha = _localizar_bolha_uc(page)
        n2 = qtd_msgs(page)
        t2 = ultima_msg_texto(page)
        r = _clicar_sim_na_bolha(page, bolha, n2, t2)
        print(f"  [Lista UC] Resposta após Sim: {r[:120]}")
        if not r:
            return False, "ERRO: bot não confirmou UC após 'Sim'"
    return True, r


def _detectar_painel_ver_faturas(page):
    try:
        ultima_msg_in = _ultima_msg_elem(page)
        if ultima_msg_in is None:
            return False
        btn = ultima_msg_in.locator(
            'button[title="Ver faturas"], button:has-text("Ver faturas")')
        for i in range(btn.count()):
            try:
                if btn.nth(i).is_visible():
                    return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def processar_lista_faturas(page):
    print("  [Lista Faturas] Abrindo painel 'Ver faturas'...")
    clicou_ver = False
    for sel in [
        'button[title="Ver faturas"]',
        'button:has-text("Ver faturas")',
        'div[role="button"]:has-text("Ver faturas")',
    ]:
        try:
            btn_ver = page.locator(sel).last
            if btn_ver.count() > 0 and btn_ver.is_visible():
                btn_ver.click(timeout=6000)
                clicou_ver = True
                print(f"  [Lista Faturas] Clicou em 'Ver faturas' ({sel})")
                time.sleep(2)
                break
        except Exception:
            continue
    if not clicou_ver:
        print("  [Lista Faturas] Botão 'Ver faturas' não encontrado.")
        return False
    itens = _listar_itens_painel(page)
    if not itens:
        print("  [Lista Faturas] Nenhum item encontrado no painel.")
        try:
            page.keyboard.press("Escape")
            time.sleep(0.5)
        except Exception:
            pass
        return False
    print(
        f"  [Lista Faturas] {len(itens)} fatura(s) no painel. Selecionando a 1ª (mais recente)...")
    for idx, item in enumerate(itens, 1):
        try:
            texto = item.inner_text().strip().replace("\n", " | ")
            marca = "← ESCOLHIDA" if idx == 1 else ""
            print(f"    [{idx}] {texto[:100]}  {marca}")
        except Exception:
            continue
    MAX_TENTATIVAS_CLIQUE = 2
    for tentativa_clique in range(1, MAX_TENTATIVAS_CLIQUE + 1):
        print(
            f"  [Lista Faturas] Tentativa {tentativa_clique}/{MAX_TENTATIVAS_CLIQUE}...")
        if tentativa_clique > 1:
            painel_aberto = False
            try:
                modal = page.locator(
                    'div[data-testid="popup-contents"], div[data-animate-modal-popup="true"]')
                painel_aberto = modal.count() > 0
            except Exception:
                painel_aberto = False
            if not painel_aberto:
                print("  [Lista Faturas] Reabrindo painel 'Ver faturas'...")
                reabriu = False
                for sel in ['button[title="Ver faturas"]', 'button:has-text("Ver faturas")']:
                    try:
                        btn_ver = page.locator(sel).last
                        if btn_ver.count() > 0 and btn_ver.is_visible():
                            btn_ver.click(timeout=5000)
                            reabriu = True
                            time.sleep(2)
                            break
                    except Exception:
                        continue
                if reabriu:
                    itens = _listar_itens_painel(page)
                    if not itens:
                        print("  [Lista Faturas] Itens não reencontrados.")
                        break
        item_fatura1 = itens[0]
        print("  [Lista Faturas] Clicando no radio da Fatura 1...")
        if not _clicar_no_item_uc(page, item_fatura1):
            print("  [Lista Faturas] Clique no radio falhou nesta tentativa.")
            time.sleep(1)
            continue
        time.sleep(1)
        n_antes = qtd_msgs(page)
        t_antes = ultima_msg_texto(page)
        pdfs_antes = contar_pdfs(page)
        print("  [Lista Faturas] Clicando no botão verde de confirmação...")
        confirmou = _confirmar_selecao_painel(page)
        if not confirmou:
            print(
                "  [Lista Faturas] Confirmação não encontrada — tentando Escape...")
            try:
                page.keyboard.press("Escape")
                time.sleep(0.8)
            except Exception:
                pass
        r = aguardar_bot(page, n_antes, t_antes, timeout=TIMEOUT_POS_CLIQUE_PAINEL,
                         palavras_chave=["informações para pagamento", "informacoes para pagamento",
                                         "pix", ".pdf", "segunda via", "⚡", "encerrado", "fatura"])
        if contar_pdfs(page) > pdfs_antes or r:
            print("  [Lista Faturas] ✓ Fatura 1 selecionada e confirmada.")
            return True
        print(f"  [Lista Faturas] Sem reação na tentativa {tentativa_clique}.")
    print("  [Lista Faturas] Falha após todas as tentativas.")
    return False


def aguardar_pdf(page, qtd_antes, texto_antes, pdfs_antes, timeout=TIMEOUT_BOT, palavras_chave=None):
    palavras_chave = [p.lower() for p in (palavras_chave or [])]
    limite = time.time() + timeout
    pix_avisado = False
    info_pagamento_em_t = None
    fatura_digital_enviada = False
    multiplas_faturas_respondido = False
    tentativas_fatura = 0
    MAX_TENTATIVAS_FATURA = 2

    def _pdf_presente():
        if _pdf_no_dom(page):
            return True
        return contar_pdfs(page) > pdfs_antes

    def _checa_e_aciona_fatura_digital(agora):
        nonlocal fatura_digital_enviada, multiplas_faturas_respondido
        if fatura_digital_enviada:
            return False
        if info_pagamento_em_t is None:
            return False
        if (agora - info_pagamento_em_t) < TIMEOUT_ESPERA_PDF_POS_INFO:
            return False
        if _pdf_presente():
            print(
                "  [aguardar_pdf] PDF detectado na checagem extra — NÃO envia 'Fatura digital'.")
            return False
        print(
            f"  [aguardar_pdf] {TIMEOUT_ESPERA_PDF_POS_INFO}s sem PDF após 'informações para pagamento' — enviando 'Fatura digital'...")
        try:
            enviar(page, "Fatura digital")
            fatura_digital_enviada = True
            multiplas_faturas_respondido = False
            print("  [aguardar_pdf] Flag de múltiplas faturas resetada.")
            return True
        except Exception as e:
            print(f"  [aguardar_pdf] erro ao enviar 'Fatura digital': {e}")
            return False

    while time.time() < limite:
        try:
            if _pdf_presente():
                print(
                    "  [aguardar_pdf] PDF detectado (DOM/contagem) — saindo imediatamente.")
                time.sleep(0.5)
                return "__PDF_RECEBIDO__"
        except Exception as e:
            print(f"  [aguardar_pdf] Erro na checagem DOM de PDF: {e}")

        try:
            qtd_atual = qtd_msgs(page)
            texto_atual = ultima_msg_texto(page).lower()
        except Exception:
            time.sleep(0.5)
            continue

        if ("melhorar nosso serviço" in texto_atual or "melhorar nosso servico" in texto_atual):
            print("  [aguardar_pdf] Mensagem de avaliação detectada — saindo.")
            return ultima_msg_texto(page)
        if (FRASE_FALTOU_NUMERO in texto_atual or FRASE_FALTOU_NUMERO_ALT in texto_atual):
            print("  [aguardar_pdf] BOT pediu CPF/CNPJ com mais dígitos — saindo.")
            return ultima_msg_texto(page)

        if not multiplas_faturas_respondido and verificar_msg_multiplas_faturas(texto_atual):
            if tentativas_fatura >= MAX_TENTATIVAS_FATURA:
                print(
                    f"  [aguardar_pdf] {tentativas_fatura}x tentativas — desistindo.")
                return "__ERRO_BAIXAR_FATURA__"
            tentativas_fatura += 1
            print(
                f"  [aguardar_pdf] BOT pediu para escolher fatura (tentativa {tentativas_fatura}/{MAX_TENTATIVAS_FATURA}).")
            try:
                if _detectar_painel_ver_faturas(page):
                    print(
                        "  [aguardar_pdf] Painel 'Ver faturas' detectado — selecionando a 1ª...")
                    if processar_lista_faturas(page):
                        multiplas_faturas_respondido = True
                        qtd_antes = qtd_msgs(page)
                        texto_antes = ultima_msg_texto(page).lower()
                        time.sleep(1)
                        continue
                    else:
                        print("  [aguardar_pdf] processar_lista_faturas falhou.")
                else:
                    print(
                        "  [aguardar_pdf] Sem painel 'Ver faturas' — enviando '1'...")
                    n_lista = qtd_msgs(page)
                    t_lista = ultima_msg_texto(page)
                    enviar(page, "1")
                    multiplas_faturas_respondido = True
                    qtd_antes = n_lista
                    texto_antes = t_lista.lower()
                    time.sleep(1)
                    continue
                multiplas_faturas_respondido = True
                qtd_antes = qtd_msgs(page)
                texto_antes = ultima_msg_texto(page).lower()
                time.sleep(1)
                continue
            except Exception as e:
                print(f"  [aguardar_pdf] Erro ao escolher fatura: {e}")

        if info_pagamento_em_t is None and verificar_msg_info_pagamento(texto_atual):
            info_pagamento_em_t = time.time()
            print(
                f"  [aguardar_pdf] 'Informações para pagamento' detectada — cronômetro iniciado ({TIMEOUT_ESPERA_PDF_POS_INFO}s para 'Fatura digital').")
        if _checa_e_aciona_fatura_digital(time.time()):
            qtd_antes = qtd_msgs(page)
            texto_antes = ultima_msg_texto(page).lower()
            time.sleep(1)
            continue

        if verificar_mensagem_pix(texto_atual) and not pix_avisado:
            print(
                "  [aguardar_pdf] Mensagem do PIX detectada — continuando a aguardar PDF...")
            pix_avisado = True
            qtd_antes = qtd_atual
            texto_antes = texto_atual
            time.sleep(1)
            continue

        mudou = (qtd_atual > qtd_antes) or (
            texto_atual and texto_atual != texto_antes.lower())
        if mudou:
            time.sleep(1.0)
            if _pdf_presente():
                print("  [aguardar_pdf] PDF detectado após mudança — saindo.")
                return "__PDF_RECEBIDO__"
            resposta = ultima_msg_texto(page)
            resposta_lower = resposta.lower()
            if info_pagamento_em_t is None and verificar_msg_info_pagamento(resposta_lower):
                info_pagamento_em_t = time.time()
                print(
                    "  [aguardar_pdf] 'Informações para pagamento' detectada (re-check).")
                qtd_antes = qtd_msgs(page)
                texto_antes = resposta_lower
                time.sleep(0.5)
                continue
            if verificar_mensagem_pix(resposta):
                if not pix_avisado:
                    print(
                        "  [aguardar_pdf] Mensagem do PIX (re-detect) — aguardando PDF...")
                    pix_avisado = True
                qtd_antes = qtd_msgs(page)
                texto_antes = resposta_lower
                time.sleep(1)
                continue
            if not palavras_chave:
                if info_pagamento_em_t is not None and not fatura_digital_enviada:
                    qtd_antes = qtd_msgs(page)
                    texto_antes = resposta_lower
                    time.sleep(0.5)
                    continue
                return resposta
            for p in palavras_chave:
                if p in resposta_lower:
                    if info_pagamento_em_t is not None and not fatura_digital_enviada:
                        qtd_antes = qtd_msgs(page)
                        texto_antes = resposta_lower
                        time.sleep(0.5)
                        break
                    return resposta
        time.sleep(0.5)

    if _pdf_presente():
        print("  [aguardar_pdf] PDF encontrado na checagem final pós-timeout.")
        return "__PDF_RECEBIDO__"
    return ""


def _conversa_aberta(page):
    try:
        if page.locator("#main").count() == 0:
            return False
        for sel in SELETORES_CAIXA_MSG:
            if page.locator(sel).count() > 0:
                return True
        return False
    except Exception:
        return False


def _reabrir_conversa_se_necessario(page):
    if _conversa_aberta(page):
        return True
    print("  [reabrir] Conversa fechada — reabrindo...")
    try:
        page.keyboard.press("Escape")
        time.sleep(0.4)
        page.keyboard.press("Escape")
        time.sleep(0.4)
    except Exception:
        pass
    if abrir_conversa(page, CONTATO_BOT):
        time.sleep(1.5)
        return _conversa_aberta(page)
    return False


def _selecionar_opcao_por_uc(page, uc_arquivo, msg_lista):
    uc_arq = digitos(uc_arquivo)
    if len(uc_arq) < 4:
        return None
    print(f"  [P6 select] UC={uc_arq}")
    linhas = [l.strip() for l in msg_lista.split("\n") if l.strip()]
    for linha in linhas:
        if not uc_bate(uc_arquivo, linha):
            continue
        m = re.match(r"^(\d+)\s*[-–.]", linha)
        if m:
            print(
                f"  [P6 select] Linha compatível: '{linha[:80]}' → opção {m.group(1)}")
            return int(m.group(1))
        m = re.search(r"[Uu]nidade\s+(\d+)", linha)
        if m:
            print(
                f"  [P6 select] Linha compatível (Unidade N): '{linha[:80]}' → opção {m.group(1)}")
            return int(m.group(1))
        m_any = re.match(r"^(\d+)", linha)
        if m_any:
            print(
                f"  [P6 select] Linha compatível (fallback): '{linha[:80]}' → opção {m_any.group(1)}")
            return int(m_any.group(1))
    print("  [P6 select] Sem match no texto — tentando botões interativos...")
    try:
        msgs = _all_msg_recebidas(page)
        for msg in reversed(msgs):
            try:
                botoes = msg.locator('button, div[role="button"]').all()
                for idx, btn in enumerate(botoes):
                    try:
                        txt = btn.inner_text().strip()
                        if uc_bate(uc_arquivo, txt):
                            m_btn = re.match(r"^(\d+)\s*[-–.]", txt)
                            num = int(m_btn.group(1)) if m_btn else idx + 1
                            print(
                                f"  [P6 select] Botão compatível: '{txt[:60]}' → opção {num}")
                            return num
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception as e:
        print(f"  [P6 select] Erro ao ler botões: {e}")
    print("  [P6 select] Nenhuma opção compatível encontrada.")
    return None


# ──────────────────────────────────────────────────────────────────────────────
# FLUXO PRINCIPAL DE PROCESSAMENTO
# ──────────────────────────────────────────────────────────────────────────────

def processar_fluxo(page, uc_arquivo, doc, data, codvenda):
    def erro(mensagem):
        print(f"  [ERRO] {mensagem} → enviando 'Sair'")
        enviar_sair_e_aguardar_encerramento(page)
        return mensagem

    def check_interrupcao(r):
        if verificar_menu_servicos(r):
            return False, "Erro: fluxo interrompido - menu de serviços fora de hora"
        nova = tratar_pergunta_dividas(page, r)
        if verificar_menu_servicos(nova):
            return False, "Erro: fluxo interrompido - menu após dívidas"
        return True, nova

    print("\n[P1] Iniciando — validando saudação Neoenergia...")
    ok_saudacao, motivo_saudacao = validar_saudacao_neoenergia(page)
    if not ok_saudacao:
        return erro(motivo_saudacao)

    print("[P2] Enviando 'segunda via'...")
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    enviar(page, "segunda via")
    r = aguardar_bot(page, n, t, timeout=TIMEOUT_BOT,
                     palavras_chave=["cpf", "cnpj", "documento", "informe"])
    print(f"  Bot: {r[:120]}")
    if not r:
        print("[P2] Sem resposta — reenviando 'segunda via' com timeout reduzido...")
        n = qtd_msgs(page)
        t = ultima_msg_texto(page)
        enviar(page, "segunda via")
        r = aguardar_bot(page, n, t, timeout=TIMEOUT_REENVIO,
                         palavras_chave=["cpf", "cnpj", "documento", "informe"])
        print(f"  Bot após reenvio: {r[:120]}")
        if not r:
            return erro("ERRO: bot não pediu CPF após 'segunda via'")

    cont, r_ou_erro = check_interrupcao(r)
    if not cont:
        return erro(r_ou_erro)
    r = r_ou_erro
    eh_erro, obs_erro = detectar_erro_bot(r)
    if eh_erro:
        return erro(obs_erro)
    if not any(p in r.lower() for p in ["cpf", "cnpj", "documento", "informe"]):
        return erro(f"ERRO: bot não pediu CPF — resposta inesperada: {r[:80]}")

    print(f"[P3] Enviando documento: {doc}")
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    enviar(page, doc)
    r = aguardar_bot(page, n, t, timeout=TIMEOUT_BOT,
                     palavras_chave=["data", "nascimento", "inválido", "invalido",
                                     "não encontrei", "nao encontrei"])
    print(f"  Bot: {r[:120]}")
    if not r:
        return erro("ERRO: bot não respondeu após CPF/CNPJ")
    cont, r_ou_erro = check_interrupcao(r)
    if not cont:
        return erro(r_ou_erro)
    r = r_ou_erro
    eh_erro, obs_erro = detectar_erro_bot(r)
    if eh_erro:
        return erro(obs_erro)
    rl = r.lower()
    if "inválido" in rl or "invalido" in rl:
        return erro(f"CPF/CNPJ rejeitado pelo bot: {r[:80]}")
    if "não encontrei" in rl or "nao encontrei" in rl:
        return erro("Bot não encontrou o CPF/CNPJ informado")
    if "data" not in rl and "nascimento" not in rl:
        return erro(f"Resposta inesperada após CPF/CNPJ: {r[:80]}")

    print(f"[P4] Enviando data: {data}")
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    enviar(page, data)
    r = aguardar_bot(page, n, t, timeout=TIMEOUT_BOT,
                     palavras_chave=["unidade consumidora", "posso seguir",
                                     "não encontrei", "nao encontrei",
                                     "inválid", "invalido", "formato",
                                     "escolha", "selecione", "ver unidades",
                                     "nenhuma fatura", "dívida", "divida",
                                     "consultar"])
    print(f"  Bot: {r[:120]}")
    if not r:
        return erro("ERRO: bot não respondeu após a data")
    cont, r_ou_erro = check_interrupcao(r)
    if not cont:
        return erro(r_ou_erro)
    r = r_ou_erro
    eh_erro, obs_erro = detectar_erro_bot(r)
    if eh_erro:
        return erro(obs_erro)
    rl = r.lower()
    if "inválid" in rl or "invalido" in rl or "formato" in rl:
        return erro(f"Data {data} rejeitada pelo bot: {r[:80]}")
    if "não encontrei" in rl or "nao encontrei" in rl:
        return erro("Bot não encontrou UC com a data informada")

    print("[P5] Verificando UC...")
    pdfs_antes_p5 = contar_pdfs(page)
    print(f"  [P5] PDFs na conversa antes do 'Sim': {pdfs_antes_p5}")
    ha_lista = False
    try:
        ultima_msg_in = _ultima_msg_elem(page)
        if ultima_msg_in is not None:
            btn_ver = ultima_msg_in.locator(
                'button[title="Ver unidades"], button:has-text("Ver unidades")')
            visiveis = 0
            for i in range(btn_ver.count()):
                try:
                    if btn_ver.nth(i).is_visible():
                        visiveis += 1
                except Exception:
                    continue
            ha_lista = visiveis > 0
            print(
                f"  [P5] Botões 'Ver unidades' visíveis na última msg: {visiveis}")
    except Exception as e:
        print(f"  [P5] Erro ao buscar 'Ver unidades': {e}")
        ha_lista = False
    if not ha_lista:
        rl_r = r.lower()
        gatilhos_lista_uc = [
            "ver unidades", "qual é a sua unidade", "qual e a sua unidade",
            "qual unidade consumidora", "selecione a unidade",
        ]
        ha_lista = any(g in rl_r for g in gatilhos_lista_uc)
        if ha_lista:
            print(f"  [P5] Lista detectada por gatilho textual.")
    print(f"  Modo: {'LISTA' if ha_lista else 'UC única'}")
    if ha_lista:
        ok, obs_p5 = processar_lista_uc(page, uc_arquivo)
        if not ok:
            if not str(obs_p5).startswith("ERRO: painel de UC não respondeu"):
                enviar_sair_e_aguardar_encerramento(page)
            return obs_p5
    else:
        ok, obs_p5 = processar_uc_unica(page, uc_arquivo, texto_uc=r)
        if not ok:
            return erro(obs_p5)
    if isinstance(obs_p5, str):
        cont, r_ou_erro = check_interrupcao(obs_p5)
        if not cont:
            return erro(r_ou_erro)
        obs_p5 = r_ou_erro
        obs_p5_lower = str(obs_p5).lower()

        # Detectar "sem fatura" diretamente em obs_p5 (caso seja a última mensagem
        # retornada) — leitura mais robusta do MK8, mesmo resultado final do MK7.2.
        if "nenhuma fatura vencida" in obs_p5_lower or "fatura vencida ou em aberto" in obs_p5_lower:
            print("[P5→P6] Bot informou sem fatura em obs_p5 → Ativa sem segunda via.")
            enviar_sair_e_aguardar_encerramento(page)
            return "Ativa - sem segunda via para emissão"

        if "posso te ajudar" in obs_p5_lower:
            # Verificar se há mensagem recente indicando que não há fatura em aberto.
            # Isso ocorre quando o bot informa "não tem nenhuma fatura vencida ou em aberto"
            # e em seguida pergunta "Posso te ajudar com algo mais?" — a UC está ativa.
            dados_js_p5 = _ler_msgs_js(page)
            todos_p5 = dados_js_p5.get("todos", [])
            texto_recentes = "\n".join(todos_p5).lower()
            _sem_fatura = ("nenhuma fatura vencida" in texto_recentes or
                           "fatura vencida ou em aberto" in texto_recentes)
            if _sem_fatura:
                enviar_sair_e_aguardar_encerramento(page)
                return "Ativa - sem segunda via para emissão"
            return erro("ERRO: bot perguntou 'posso te ajudar' após confirmar UC")

    print("[P6/P7] Verificando seleção de fatura / aguardando PDF...")
    # Verificar antecipadamente se bot já informou sem fatura (evita timeout desnecessário)
    dados_js_entrada_p67 = _ler_msgs_js(page)
    todos_entrada_p67 = "\n".join(dados_js_entrada_p67.get("todos", [])).lower()
    if "nenhuma fatura vencida" in todos_entrada_p67 or "fatura vencida ou em aberto" in todos_entrada_p67:
        print("[P6/P7] Bot já informou sem fatura ao entrar em P6/P7 → Ativa sem segunda via.")
        enviar_sair_e_aguardar_encerramento(page)
        return "Ativa - sem segunda via para emissão"
    n = qtd_msgs(page)
    t = ultima_msg_texto(page)
    pdfs_antes = pdfs_antes_p5
    time.sleep(1.5)
    r = ""

    ha_painel_p6 = False
    try:
        ultima_msg_in_p6 = _ultima_msg_elem(page)
        if ultima_msg_in_p6 is not None:
            btn_ver_p6 = ultima_msg_in_p6.locator(
                'button[title="Ver unidades"], button:has-text("Ver unidades")')
            visiveis = 0
            for i in range(btn_ver_p6.count()):
                try:
                    if btn_ver_p6.nth(i).is_visible():
                        visiveis += 1
                except Exception:
                    continue
            ha_painel_p6 = visiveis > 0
    except Exception:
        ha_painel_p6 = False

    if ha_painel_p6:
        print("[P6] Painel 'Ver unidades' detectado.")
        ok_p6, obs_p6 = processar_lista_uc(page, uc_arquivo)
        if not ok_p6:
            return erro(f"ERRO P6 painel Ver unidades: {obs_p6}")
        n = qtd_msgs(page)
        t = ultima_msg_texto(page)
        pdfs_antes = contar_pdfs(page)
        r = aguardar_pdf(page, n, t, pdfs_antes, timeout=TIMEOUT_BOT,
                         palavras_chave=["pdf", ".pdf", "fatura", "⚡", "encerrado", "nenhuma fatura"])

    elif existe_pdf_recente(page, pdfs_antes):
        print("[P7] PDF disponível diretamente.")
        r = "__PDF_RECEBIDO__"

    else:
        r = aguardar_pdf(page, n, t, pdfs_antes, timeout=TIMEOUT_BOT,
                         palavras_chave=[".pdf", "segunda via -", "nenhuma fatura", "⚡",
                                         "ver unidades", "ver opções", "ver opcoes",
                                         "dívida", "divida", "consultar",
                                         "não entendi", "nao entendi"])
        if r and r != "__PDF_RECEBIDO__":
            r = tratar_pergunta_dividas(page, r)
            if verificar_menu_servicos(r):
                return erro("Erro: fluxo interrompido - menu antes do PDF")
            if existe_pdf_recente(page, pdfs_antes):
                r = "__PDF_RECEBIDO__"

        try:
            ha_tardio = False
            ultima_msg_in_tardio = _ultima_msg_elem(page)
            if ultima_msg_in_tardio is not None:
                btn_ver_tardio = ultima_msg_in_tardio.locator(
                    'button[title="Ver unidades"], button:has-text("Ver unidades")')
                for i in range(btn_ver_tardio.count()):
                    try:
                        if btn_ver_tardio.nth(i).is_visible():
                            ha_tardio = True
                            break
                    except Exception:
                        continue
            if ha_tardio and r != "__PDF_RECEBIDO__":
                print("[P6] Painel 'Ver unidades' detectado (tardio).")
                ok_p6, obs_p6 = processar_lista_uc(page, uc_arquivo)
                if not ok_p6:
                    return erro(f"ERRO P6 painel Ver unidades (tardio): {obs_p6}")
                n = qtd_msgs(page)
                t = ultima_msg_texto(page)
                pdfs_antes = contar_pdfs(page)
                r = aguardar_pdf(page, n, t, pdfs_antes, timeout=TIMEOUT_BOT,
                                 palavras_chave=["pdf", ".pdf", "fatura", "⚡", "encerrado", "nenhuma fatura"])
        except Exception:
            pass

        if r not in ("__PDF_RECEBIDO__", "") and not existe_pdf_recente(page, pdfs_antes):
            time.sleep(1.5)
            if existe_pdf_recente(page, pdfs_antes):
                print("[P6/P7] PDF detectado em re-check — pulando seleção.")
                r = "__PDF_RECEBIDO__"
            else:
                msg_lista = ultima_msg_texto(page)
                print(f"[P6] Mensagem do bot: {msg_lista[:200]}")
                tem_uc_mascarada = bool(re.search(r"\d+\*+\d+", msg_lista))
                if tem_uc_mascarada:
                    numero_escolhido = _selecionar_opcao_por_uc(
                        page, uc_arquivo, msg_lista)
                else:
                    print("[P6] Texto não contém UC mascarada — não é lista de UCs.")
                    numero_escolhido = None
                if numero_escolhido is None:
                    print(
                        "[P6/P7] Sem lista de UCs identificável. Aguardando PDF...")
                    n = qtd_msgs(page)
                    t = ultima_msg_texto(page)
                    pdfs_antes_n = contar_pdfs(page)
                    r = aguardar_pdf(page, n, t, pdfs_antes_n, timeout=TIMEOUT_BOT,
                                     palavras_chave=["pdf", ".pdf", "fatura", "⚡", "encerrado", "nenhuma fatura"])
                    if r and r != "__PDF_RECEBIDO__":
                        r = tratar_pergunta_dividas(page, r)
                        if existe_pdf_recente(page, pdfs_antes):
                            r = "__PDF_RECEBIDO__"
                else:
                    print(f"[P6] Enviando opção '{numero_escolhido}'...")
                    n = qtd_msgs(page)
                    t = ultima_msg_texto(page)
                    pdfs_antes = contar_pdfs(page)
                    enviar(page, str(numero_escolhido))
                    r = aguardar_pdf(page, n, t, pdfs_antes, timeout=TIMEOUT_BOT,
                                     palavras_chave=["fatura", "pdf", ".pdf", "sair",
                                                     "encerrar", "dica", "⚡",
                                                     "obrigado", "menu", "nenhuma fatura",
                                                     "dívida", "divida", "consultar"])
                    if r and r != "__PDF_RECEBIDO__":
                        r = tratar_pergunta_dividas(page, r)
                        if existe_pdf_recente(page, pdfs_antes):
                            r = "__PDF_RECEBIDO__"

    if r == "__ERRO_BAIXAR_FATURA__":
        print("[P7] Erro ao selecionar fatura — Sair + retry.")
        try:
            enviar_sair_e_aguardar_encerramento(page)
        except Exception:
            pass
        return "Erro ao baixar fatura"

    if r == "__PDF_RECEBIDO__" or existe_pdf_recente(page, pdfs_antes):
        codvenda_str = str(codvenda or "").strip()
        nome_sanitizado = re.sub(r"[^\w\-.]", "_", codvenda_str).strip("._")
        if not nome_sanitizado:
            print(f"  [P7] AVISO: codvenda vazio/inválido (raw='{codvenda}')")
            nome_sanitizado = f"sem_codvenda_{datetime.now().strftime('%H%M%S')}"
        nome = f"{nome_sanitizado}.pdf"
        destino = (PASTA_FATURAS / nome).resolve()
        sufixo = 2
        while destino.exists():
            nome = f"{nome_sanitizado}_{sufixo}.pdf"
            destino = (PASTA_FATURAS / nome).resolve()
            sufixo += 1
            if sufixo > 100:
                break
        print(f"[P7] PDF localizado → baixando imediatamente: {destino}")
        # Até 2 tentativas de download. A UC já está confirmada e o PDF apareceu,
        # então o resultado será Ativa mesmo que o download em si não conclua.
        baixou = False
        for tentativa_dl in range(1, 3):
            print(f"[P7] Tentativa de download {tentativa_dl}/2...")
            try:
                if baixar_pdf(page, destino):
                    baixou = True
                    break
            except Exception as e:
                print(f"[P7] Erro na tentativa {tentativa_dl} de download: {e}")
            time.sleep(2)
        if baixou:
            if destino.exists() and destino.stat().st_size > 0:
                tamanho_kb = destino.stat().st_size // 1024
                print(f"[P7] ✓ Fatura salva como '{nome}' ({tamanho_kb} KB)")
                obs_dl = f"validada e baixada - Faturas/{nome}"
            else:
                obs_dl = f"PDF recebido - arquivo não foi criado em {destino}"
                print(
                    f"  AVISO: baixar_pdf retornou True mas {destino} não existe ou está vazio.")
        else:
            # PDF apareceu na tela mas o download falhou nas 2 tentativas → Ativa
            obs_dl = f"PDF recebido - download falhou para {codvenda_str}"
        print("[P8] Enviando 'Sair' após download...")
        enviar_sair_e_aguardar_encerramento(page)
        return obs_dl

    # Nesta altura a UC JÁ foi confirmada (passou por P5). Se o bot exibe a tela de
    # avaliação ("melhorar nosso serviço"), o atendimento foi encerrado normalmente
    # sem PDF — NÃO é erro. Mandamos 'Sair' para reiniciar o bot e marcamos Ativa.
    if _eh_avaliacao(r):
        print("  [P6/P7] Avaliação após UC confirmada → Ativa (sem segunda via). Reiniciando bot...")
        enviar_sair_e_aguardar_encerramento(page)
        return "Ativa - sem segunda via para emissão"
    eh_erro, obs_erro = detectar_erro_bot(r)
    if eh_erro:
        return erro(obs_erro)

    print("[P8] Enviando 'Sair'...")
    encerrado = enviar_sair_e_aguardar_encerramento(page)
    if encerrado:
        return "validada"
    return "Atendimento finalizado (sem confirmação de encerramento)"


# ──────────────────────────────────────────────────────────────────────────────
# PROCESSAMENTO DE LINHA (wrapper)
# ──────────────────────────────────────────────────────────────────────────────

def processar_linha(page, row, col):
    codvenda_raw = row.get(col["codvenda"]) if col["codvenda"] else None
    uc_raw = row.get(col["uc"])
    doc_raw = row.get(col["doc"])
    codvenda = str(codvenda_raw).strip() if str(
        codvenda_raw) not in ("None", "nan", "") else "sem_cod"
    uc = digitos(uc_raw)
    doc = digitos(doc_raw)
    data = None
    for col_data in col["datas"]:
        data = parse_data(row.get(col_data))
        if data:
            break
    print("\n" + "─" * 80)
    print(f"  CODVENDA={codvenda}  UC={uc}  DOC={doc}  DATA={data}")
    # Linhas puladas (dado inválido na planilha): não dá para saber se a UC é
    # Ativa/Inativa sem processar, então marcamos "Validar Documentos".
    if not uc:
        return "Validar Documentos", "pulado", ""
    if not doc or len(doc) not in (11, 14):
        return "Validar Documentos", "pulado", ""
    if not data:
        return "Validar Documentos", "pulado", ""

    busca = ""  # preenchido só quando um localizador/elemento não foi encontrado
    try:
        obs = processar_fluxo(page, uc, doc, data, codvenda)
    except Exception as e:
        obs = f"ERRO inesperado: {e}"
        # Erros de localizador alimentam a coluna 'Busca' com o que se procurava:
        #  · "Não encontrei a caixa de mensagem. Seletores buscados: ..."
        #  · "Locator.hover Timeout ..." (elemento de PDF não respondeu ao hover)
        msg_exc = str(e)
        msg_lower = msg_exc.lower()
        if "não encontrei a caixa de mensagem" in msg_lower or \
                "nao encontrei a caixa de mensagem" in msg_lower or \
                "hover" in msg_lower or "timeout" in msg_lower or \
                "locator" in msg_lower:
            busca = msg_exc
        try:
            enviar_sair_e_aguardar_encerramento(page)
        except Exception:
            pass
    print(f"  → {obs}")
    obs_lower = obs.lower()

    # ── Classificação de TIPO (controla contadores e fila de retry) ──────────
    # As strings "validada...", "PDF recebido...", "Ativa - sem segunda via..."
    # vêm de processar_fluxo; aqui são traduzidas para o Resultado canônico.
    pdf_baixado = obs.startswith("validada e baixada")
    # "dados desatualizados": a UC EXISTE, apenas o cadastro está desatualizado
    # → conforme regra, marcamos como Ativa (não Inativa).
    dados_desatualizados = "dados desatualizados" in obs_lower
    if pdf_baixado or obs.startswith("validada") or obs.startswith("PDF recebido") or \
            obs == "Ativa - sem segunda via para emissão":
        tipo = "sucesso"
    elif dados_desatualizados:
        tipo = "sucesso"  # UC ativa, cadastro desatualizado → Ativa
    elif "inativa" in obs_lower:
        tipo = "inativa"
    # "UC não confere..." e "mais de uma UC - nenhuma conferiu..." agora são Inativa
    elif obs.startswith("UC não confere"):
        tipo = "inativa"
    elif obs.startswith("mais de uma UC - nenhuma conferiu"):
        tipo = "inativa"
    elif obs.startswith("Erro CPF"):
        tipo = "falha"
    elif obs.startswith("mais de uma UC"):
        # Outros casos de "mais de uma UC" (botão/itens não encontrados) → falha
        tipo = "falha"
    elif eh_erro_de_retry(obs):
        tipo = "retry"
    else:
        tipo = "falha"

    # ── Tradução para a coluna RESULTADO canônica ────────────────────────────
    if tipo == "pulado":
        resultado = obs  # "Validar Documentos"
    elif tipo == "sucesso":
        # PDF efetivamente salvo → "Ativa e Fatura baixada"; senão só "Ativa"
        resultado = RES_ATIVA_FATURA if pdf_baixado else RES_ATIVA
    elif tipo == "inativa":
        resultado = RES_INATIVA
    elif obs_lower == "sem fatura em aberto":
        resultado = RES_ATIVA
    else:
        # falha ou retry: o resultado provisório é o próprio texto do erro.
        # No retry, a fila reprocessa e sobrescreve com o resultado final.
        resultado = obs

    return resultado, tipo, busca


# ──────────────────────────────────────────────────────────────────────────────
# ENTRADA DE DADOS
# ──────────────────────────────────────────────────────────────────────────────

def detectar_e_carregar_arquivo():
    candidatos = [
        PASTA_PROJETO / "Elektro_1.xlsx",
        PASTA_PROJETO / "Elektro_1.xls",
        PASTA_PROJETO / "Elektro_1.csv",
        PASTA_PROJETO / "Elektro.xlsx",
        PASTA_PROJETO / "Elektro.csv",
        PASTA_PROJETO / "Elektro 1.xlsx",
        PASTA_PROJETO / "Elektro 1.csv",
        ARQUIVO_ENTRADA,
    ]
    arquivo = next((c for c in candidatos if c.exists()), None)
    if arquivo is None:
        sys.exit(
            f"ERRO: nenhum arquivo de entrada encontrado em: {PASTA_PROJETO}")
    print(f"Arquivo de entrada: {arquivo}")
    sufixo = arquivo.suffix.lower()
    try:
        if sufixo in (".xlsx", ".xls"):
            df = pd.read_excel(arquivo, dtype=str)
        elif sufixo == ".csv":
            df = None
            # Testa múltiplas codificações para cada separador
            for sep in [";", ",", "\t"]:
                for enc in ["utf-8", "latin-1", "cp1252", "utf-8-sig"]:
                    try:
                        tmp = pd.read_csv(arquivo, dtype=str,
                                          sep=sep, encoding=enc)
                        if tmp.shape[1] > 1:
                            df = tmp
                            break
                    except Exception:
                        continue
                if df is not None:
                    break
            if df is None:
                sys.exit(
                    "ERRO: não consegui ler o CSV com nenhum separador (; , TAB).")
        else:
            sys.exit(f"ERRO: formato não suportado: {sufixo}")
    except Exception as e:
        sys.exit(f"ERRO ao abrir o arquivo: {e}")
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df, arquivo


def mapear_colunas(df):
    """
    Mapeia colunas do df para nomes canônicos.
    Aceita variações com/sem acento usando _remover_acentos().
    """
    cols_originais = df.columns.tolist()
    # Mapeamento: nome_normalizado → nome_original
    norm_para_orig = {_remover_acentos(c): c for c in cols_originais}

    def achar(*nomes):
        for nome in nomes:
            # Tenta nome exato primeiro
            if nome in cols_originais:
                return nome
            # Tenta sem acento
            norm = _remover_acentos(nome)
            if norm in norm_para_orig:
                return norm_para_orig[norm]
        return None

    col = dict(
        codvenda=achar("codvenda", "cod venda",
                       "codigo venda", "código venda"),
        uc=achar("unidade consumidora", "uc", "instalacao", "instalação"),
        doc=achar("cpf/cnpj", "cpf / cnpj", "cpf", "cnpj", "documento"),
        # Apenas duas colunas de resultado: Resultado (sempre) e Busca (só em
        # erro de localizador). As antigas obs/validacao/ativas/inativas/fatura
        # foram unificadas em "resultado".
        resultado=achar("resultado") or COL_RESULTADO,
        busca=achar("busca") or COL_BUSCA,
    )
    col["datas"] = [c for c in cols_originais if "data" in _remover_acentos(c)]

    if col["uc"] is None or col["doc"] is None:
        sys.exit(
            f"ERRO: cabeçalho não reconhecido.\nColunas: {cols_originais}\n"
            f"Necessário: 'uc' (ou 'instalacao') e 'cpf/cnpj' (ou 'cpf'/'cnpj')."
        )
    return col


# ──────────────────────────────────────────────────────────────────────────────
# WORKBOOK — CRIAÇÃO E SALVAMENTO (CORRIGIDOS)
# ──────────────────────────────────────────────────────────────────────────────

def garantir_colunas_resultado(df, col):
    """
    Garante que as colunas 'Resultado' e 'Busca' existam no df ANTES de
    qualquer split/cópia. Retorna df (pode ser o mesmo objeto modificado).
    Chamada ANTES de criar df_proc.
    """
    for chave in ("resultado", "busca"):
        nome = col[chave]
        if nome not in df.columns:
            df[nome] = ""
    return df


def criar_workbook_do_zero(df):
    """
    Cria um workbook openpyxl a partir do df.
    Não depende de nenhum arquivo existente no disco.
    Retorna (wb, ws, col_map) onde col_map é {nome_coluna: índice_1based}.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    colunas = df.columns.tolist()
    col_map = {}

    # Cabeçalho
    for c_idx, nome in enumerate(colunas, 1):
        ws.cell(row=1, column=c_idx, value=nome)
        col_map[nome] = c_idx

    # Dados iniciais (copia o df inteiro com valores atuais)
    for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row_data, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=val)
            except Exception as e:
                print(
                    f"  AVISO criar_workbook: célula ({r_idx},{c_idx}) erro: {e}")

    return wb, ws, col_map


def gravar_resultado_linha(ws, wb_row, col_resultado_idx, resultado,
                           col_busca_idx=None, busca=""):
    """
    Grava 'Resultado' e (opcionalmente) 'Busca' na linha wb_row.
    Gravado UMA vez por UC; no retry, a fila chama de novo e SOBRESCREVE
    a mesma linha com o resultado da última tentativa.
    Inclui try/except individual para evitar silêncio em erros.
    """
    try:
        ws.cell(row=wb_row, column=col_resultado_idx,
                value=str(resultado) if resultado else None)
    except Exception as e:
        print(f"  ERRO ao gravar 'resultado' na linha {wb_row}: {e}")
    if col_busca_idx is not None:
        try:
            # Sobrescreve sempre (inclusive limpando se busca vazia no retry bem-sucedido)
            ws.cell(row=wb_row, column=col_busca_idx,
                    value=str(busca) if busca else None)
        except Exception as e:
            print(f"  ERRO ao gravar 'busca' na linha {wb_row}: {e}")


def salvar_planilha(wb, caminho):
    """
    Salva o workbook no caminho indicado.
    Se o arquivo estiver bloqueado (aberto no Excel), tenta um nome alternativo
    e imprime aviso claro.
    Retorna o Path onde o arquivo foi efetivamente salvo.
    """
    caminho = Path(caminho)
    try:
        caminho.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(caminho))
        print(f"  → Salvo: {caminho}")
        return caminho
    except PermissionError:
        alt = caminho.parent / \
            f"Elektro_resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(
            f"\n  ⚠️  AVISO: '{caminho.name}' está bloqueado (aberto no Excel?).")
        print(f"  ⚠️  Salvando em arquivo alternativo: {alt.name}\n")
        try:
            wb.save(str(alt))
            print(f"  → Salvo (alternativo): {alt}")
            return alt
        except Exception as e2:
            print(
                f"  ERRO CRÍTICO: não foi possível salvar em nenhum local: {e2}")
            return None
    except Exception as e:
        print(f"  ERRO ao salvar planilha: {e}")
        return None


def formatar_planilha_final(ws):
    """
    Aplica a formatação visual ao worksheet (chamada só no arquivo FINAL):
      · Cabeçalho com fundo verde, texto branco em negrito, centralizado
      · Demais células centralizadas, com fundo cinza claro
      · Altura de todas as linhas = 22
      · Largura de coluna ajustada ao conteúdo
    Falhas de formatação não devem impedir o salvamento → tudo em try/except.
    """
    try:
        VERDE = PatternFill("solid", fgColor="2E7D32")     # cabeçalho
        CINZA = PatternFill("solid", fgColor="F2F2F2")     # fundo das células
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fonte_cab = Font(bold=True, color="FFFFFF")

        max_row = ws.max_row
        max_col = ws.max_column

        # Cabeçalho (linha 1)
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = VERDE
            cell.font = fonte_cab
            cell.alignment = centro
        ws.row_dimensions[1].height = 22

        # Corpo
        for r in range(2, max_row + 1):
            ws.row_dimensions[r].height = 22
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = CINZA
                cell.alignment = centro

        # Largura das colunas pelo maior conteúdo (limitada para não estourar)
        for c in range(1, max_col + 1):
            maior = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    maior = max(maior, len(str(v)))
            largura = min(max(maior + 2, 12), 60)
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = largura
    except Exception as e:
        print(f"  AVISO: falha ao formatar a planilha (salvando assim mesmo): {e}")


def nome_arquivo_resultado_final(qtd_ativas):
    data_hoje = datetime.now().strftime("%d-%m-%Y")
    return PASTA_RESULTADO / f"Elektro_Resultado_{data_hoje}_{qtd_ativas}_ativas.xlsx"


# ──────────────────────────────────────────────────────────────────────────────
# ENVIO DE RESUMO AO GRUPO
# ──────────────────────────────────────────────────────────────────────────────

def enviar_resumo_grupo(page, total, sucesso, pulado, falha, inativa, duracao_segundos):
    """
    Envia resumo de execução ao grupo GOC.
    Inclui tempo total de execução formatado.
    """
    print(f"\nEnviando resumo para o grupo '{GRUPO_NOTIFICACAO}'...")
    busca, info_sel = _localizar_caixa_busca(page)
    if busca is None:
        print(f"  AVISO: {info_sel} Resumo não enviado.")
        return
    try:
        _digitar_na_busca(page, busca, GRUPO_NOTIFICACAO)
    except Exception as e:
        print(f"  AVISO: erro ao digitar nome do grupo: {e}")
        return
    abriu = False
    try:
        page.locator(f'span[title="{GRUPO_NOTIFICACAO}"]').first.click(
            timeout=10_000)
        time.sleep(2)
        abriu = True
        print(f"  Grupo '{GRUPO_NOTIFICACAO}' aberto.")
    except Exception:
        try:
            page.keyboard.press("Enter")
            time.sleep(3)
            abriu = True
            print("  Grupo aberto via ENTER.")
        except Exception as e:
            print(f"  AVISO: não consegui abrir o grupo: {e}")
            return
    if not abriu:
        return

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    tempo_str = formatar_duracao(duracao_segundos)

    resumo = (
        f"📊 *Resumo Elektro (Neoenergia) — {agora}*\n"
        f"──────────────────────\n"
        f"✅ Ativas              : {sucesso}\n"
        f"🚫 Inativas (cadastro) : {inativa}\n"
        f"🔁 Retry               : {pulado}\n"
        f"❌ Erros               : {falha}\n"
        f"📦 Total processado    : {total}\n"
        f"⏱️  Tempo de execução   : {tempo_str}\n"
        f"──────────────────────\n"
        f"Resultado salvo em: Elektro_resultado.xlsx"
    )
    try:
        box = caixa_msg(page)
        box.click()
        time.sleep(0.3)
        for linha in resumo.split("\n"):
            box.type(linha, delay=10)
            page.keyboard.press("Shift+Enter")
        page.keyboard.press("Backspace")
        page.keyboard.press("Enter")
        time.sleep(1)
        print("  Resumo enviado ao grupo.")
    except Exception as e:
        print(f"  AVISO: erro ao enviar resumo: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    tempo_inicio = time.time()  # ← Marca início para medir duração total

    df, arquivo_entrada = detectar_e_carregar_arquivo()
    col = mapear_colunas(df)

    # ── CORREÇÃO Bug 4: garante colunas obs/validacao ANTES de criar df_proc ──
    df = garantir_colunas_resultado(df, col)

    ini = max(0, LINHA_INICIAL - 1)
    fim = LINHA_FINAL or len(df)

    # df_proc agora já tem as colunas obs/validacao garantidas
    df_proc = df.iloc[ini:fim].copy().reset_index(drop=True)

    # ── CORREÇÃO Bug 1+3: workbook criado do zero, sem depender de arquivo ───
    wb, ws, col_map = criar_workbook_do_zero(df)

    # Índices de coluna para gravação (1-based, referenciados ao df inteiro)
    col_resultado_idx = col_map[col["resultado"]]
    col_busca_idx     = col_map[col["busca"]]

    # ── CORREÇÃO Bug 1: wb_row agora usa ini para corrigir o offset ──────────
    # df_proc.iterrows() retorna df_idx de 0..len(df_proc)-1 (reset_index)
    # A linha real no workbook = cabeçalho(1) + ini + df_idx + 1
    #   = df_idx + ini + 2
    # Exemplo: LINHA_INICIAL=5 → ini=4 → df_idx=0 → wb_row=4+0+2=6 ✓
    #          LINHA_INICIAL=1 → ini=0 → df_idx=0 → wb_row=0+0+2=2 ✓

    try:
        PASTA_FATURAS.mkdir(parents=True, exist_ok=True)
        teste = PASTA_FATURAS / ".write_test"
        teste.write_text("ok", encoding="utf-8")
        teste.unlink()
        print(f"Pasta de faturas OK: {PASTA_FATURAS.resolve()}")
    except Exception as e:
        print(f"AVISO: problema com a pasta de faturas ({PASTA_FATURAS}): {e}")
    try:
        PASTA_RESULTADO.mkdir(parents=True, exist_ok=True)
        print(f"Pasta de resultado OK: {PASTA_RESULTADO.resolve()}")
    except Exception as e:
        print(
            f"AVISO: problema com a pasta de resultado ({PASTA_RESULTADO}): {e}")

    print("=" * 80)
    print("AUTOMAÇÃO ELEKTRO (NEOENERGIA) — SEGUNDA VIA VIA WHATSAPP")
    print("ElektroSWMKEight")
    print("=" * 80)
    print(f"Entrada   : {arquivo_entrada}")
    print(f"Resultado : {PASTA_RESULTADO}")
    print(f"Faturas   : {PASTA_FATURAS}")
    print(f"Linhas    : {len(df_proc)} (de {ini+1} até {ini+len(df_proc)})")
    print("=" * 80)

    with sync_playwright() as p:
        print("Abrindo navegador...")
        try:
            ctx = p.chromium.launch_persistent_context(
                user_data_dir=str(PASTA_PERFIL), channel="chrome",
                headless=False, accept_downloads=True,
                args=["--start-maximized"], no_viewport=True)
        except Exception as e:
            print(f"Chrome não disponível ({e}). Usando Chromium embutido...")
            ctx = p.chromium.launch_persistent_context(
                user_data_dir=str(PASTA_PERFIL), headless=False,
                accept_downloads=True, args=["--start-maximized"], no_viewport=True)

        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.goto("https://web.whatsapp.com/", timeout=120_000)

        if not aguardar_whatsapp_carregar(page):
            input("Pressione ENTER para fechar...")
            ctx.close()
            return
        if not abrir_conversa(page, CONTATO_BOT):
            print("ERRO: não consegui abrir a conversa.")
            input("Pressione ENTER para fechar...")
            ctx.close()
            return

        print(f"\nConversa '{CONTATO_BOT}' aberta. Processando...\n")

        # ── DIAGNÓSTICO: verifica leitura de mensagens ───────────────────────
        print("[DIAGNÓSTICO] Testando leitura de mensagens recebidas via JS...")
        _diag = _ler_msgs_js(page)
        print(f"[DIAGNÓSTICO] Mensagens detectadas: {_diag['count']}")
        if _diag["count"] == 0:
            print("[DIAGNÓSTICO] AVISO: nenhuma mensagem localizada pelo JS.")
            print("[DIAGNÓSTICO] Verifique se a conversa está aberta e visível.")
        else:
            print(f"[DIAGNÓSTICO] Última mensagem: {_diag['ultimo'][:80]!r}")
        # ─────────────────────────────────────────────────────────────────────

        try:
            ultima_inicial = ultima_msg_texto(page)
            if ultima_inicial and not verificar_encerramento(ultima_inicial):
                print(
                    f"[INÍCIO] Última mensagem não é de encerramento: '{ultima_inicial[:80]}'")
                print("[INÍCIO] Enviando 'Sair' para garantir estado limpo...")
                enviar_sair_e_aguardar_encerramento(page)
            else:
                print("[INÍCIO] Estado já limpo (encerramento ou conversa vazia).")
        except Exception as e:
            print(f"[INÍCIO] AVISO: erro na limpeza inicial: {e}")

        total = sucesso = pulado = falha = inativa = 0
        fila_retry = []

        for df_idx, row in df_proc.iterrows():
            total += 1

            resultado, tipo, busca = processar_linha(page, row, col)

            # ── CORREÇÃO Bug 1: wb_row corrigido com offset ini ──────────────
            wb_row = ini + df_idx + 2

            # Grava UMA vez por UC (coluna Resultado + Busca quando houver).
            gravar_resultado_linha(
                ws, wb_row, col_resultado_idx, resultado,
                col_busca_idx=col_busca_idx, busca=busca)
            salvar_planilha(wb, ARQUIVO_RESULTADO_TMP)

            if tipo == "pulado":
                pulado += 1
            elif tipo == "sucesso":
                sucesso += 1
            elif tipo == "inativa":
                inativa += 1
            elif tipo == "retry":
                falha += 1
                fila_retry.append({
                    "df_idx": df_idx,
                    "row": row,
                    "wb_row": wb_row,
                })
            else:
                falha += 1

            print(f"\n  📋 Processadas {total} de {len(df_proc)}")

            time.sleep(PAUSA_ENTRE_ITENS)

        # ── RETRY ─────────────────────────────────────────────────────────────
        if fila_retry:
            print("\n" + "=" * 80)
            print(
                f"RETRY: {len(fila_retry)} linha(s) marcada(s) para nova tentativa.")
            print("=" * 80)
            retry_ok = retry_fail = 0
            for item in fila_retry:
                row = item["row"]
                wb_row_orig = item["wb_row"]
                print(f"\n[RETRY] Reprocessando linha (wb_row={wb_row_orig})...")
                # Reprocessa; só o resultado da ÚLTIMA tentativa é gravado.
                resultado_r, tipo_r, busca_r = processar_linha(page, row, col)
                if tipo_r == "sucesso":
                    retry_ok += 1
                    falha -= 1
                    sucesso += 1
                elif tipo_r == "inativa":
                    retry_ok += 1
                    falha -= 1
                    inativa += 1
                else:
                    retry_fail += 1

                # Sobrescreve a MESMA linha com o resultado final (e Busca, se houver).
                gravar_resultado_linha(
                    ws, wb_row_orig, col_resultado_idx, resultado_r,
                    col_busca_idx=col_busca_idx, busca=busca_r)
                salvar_planilha(wb, ARQUIVO_RESULTADO_TMP)
                time.sleep(PAUSA_ENTRE_ITENS)
            print(f"RETRY: {retry_ok} sucesso(s), {retry_fail} falha(s).")

        # ── Salvar arquivo final ───────────────────────────────────────────────
        arquivo_final = nome_arquivo_resultado_final(sucesso)
        print(f"\nSalvando arquivo final: {arquivo_final}")
        # Formatação visual aplicada SOMENTE no arquivo final (não no TMP de progresso)
        formatar_planilha_final(ws)
        caminho_salvo = salvar_planilha(wb, arquivo_final)

        # Remove temporário somente se o final foi salvo com sucesso
        try:
            if caminho_salvo and caminho_salvo.exists() and ARQUIVO_RESULTADO_TMP.exists():
                ARQUIVO_RESULTADO_TMP.unlink()
                print(
                    f"  → Removido arquivo temporário: {ARQUIVO_RESULTADO_TMP.name}")
        except Exception as e:
            print(f"  → AVISO: não foi possível remover o temporário: {e}")

        # ── Relatório final ────────────────────────────────────────────────────
        duracao_total = time.time() - tempo_inicio
        tempo_str = formatar_duracao(duracao_total)

        print("\n" + "=" * 80)
        print("FIM — ElektroSWMKEight")
        print("=" * 80)
        print(f"Total          : {total}")
        print(f"Ativas         : {sucesso}")
        print(f"Inativas       : {inativa}")
        print(f"Puladas        : {pulado}")
        print(f"Erros          : {falha}")
        print(f"Tempo total    : {tempo_str}")
        print(
            f"Arquivo        : {caminho_salvo.name if caminho_salvo else '(não salvo)'}")
        print("=" * 80)

        # ── Resumo no grupo (com tempo de execução) ───────────────────────────
        try:
            enviar_resumo_grupo(page, total, sucesso, pulado,
                                falha, inativa, duracao_total)
        except Exception as e:
            print(f"AVISO: não foi possível enviar resumo ao grupo: {e}")

        input("Pressione ENTER para fechar o navegador...")
        ctx.close()


if __name__ == "__main__":
    main()
